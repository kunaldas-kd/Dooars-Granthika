"""
transactions/views.py

Tenant resolution:  request.user.library  (OneToOneField on accounts.Library)

books.Book and members.Member are scoped by  owner = FK(User) —
the library admin's User object.  We resolve this as  library.user.

Transaction / Fine / MissingBook are scoped by  library = FK(Library)
and queried via  Model.objects.for_library(library).

═══════════════════════════════════════════════════════════════════════════════
BUSINESS RULES
═══════════════════════════════════════════════════════════════════════════════

1. ISSUE BOOK
   ✔ Member must be active
   ✔ Book must have available copies
   ✔ Member must not have exceeded their borrow limit
   ✔ Member must have no pending (unpaid) fines

2. RETURN BOOK
   ✔ Transaction must belong to this library (tenant check)
   ✔ Book must not already be returned
   ✔ If return_date > due_date → overdue Fine upserted automatically
   ✔ If book is damaged → damage Fine upserted automatically

3. RENEW BOOK
   ✔ Member must be active
   ✔ Renewal count must be below library limit
   ✔ No unpaid Fine rows on this transaction (Fine table is authoritative)
   ✔ If renewal_date > due_date → overdue Fine created BEFORE extending due_date

4. FINE APPLIED WHEN
   ✔ Late return    → Fine(TYPE_OVERDUE)  in return_book
   ✔ Late renewal   → Fine(TYPE_OVERDUE)  in renew_book
   ✔ Book lost      → Fine(TYPE_LOST)     in mark_lost / add_penalty
   ✔ Book damaged   → Fine(TYPE_DAMAGE)   in return_book

═══════════════════════════════════════════════════════════════════════════════

Overdue sync strategy
─────────────────────
A background daemon thread (fine_sync.py) runs
Transaction.sync_overdue_for_library() every FINE_SYNC_INTERVAL seconds.
Views call _sync_overdue_if_stale() which is a no-op when the background
thread has already synced within STALE_THRESHOLD_SECONDS.

Double-fine prevention
──────────────────────
fine_sync.py uses get_or_create keyed on (library, transaction, fine_type).
return_book now uses the same get_or_create pattern so whichever path runs
first wins and the second path only updates the existing unpaid row instead
of inserting a duplicate.
"""

from __future__ import annotations

from datetime import date, timedelta
from decimal import Decimal
import time

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction as db_transaction
from django.db.models import Count, Q, Sum, DecimalField as _DF
from django.http import (
    Http404,
    HttpResponse,
    JsonResponse,
)
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse

from finance.models import Fine
from .forms import (
    AddPenaltyForm,
    IssueBookForm,
    MarkFinePaidForm,
    MarkLostForm,
    ReturnBookForm,
)
from .models import MissingBook, Transaction


# ─────────────────────────────────────────────────────────────────────────────
# Email service
# ─────────────────────────────────────────────────────────────────────────────

try:
    from core import email_service as _email_svc
    _EMAIL_AVAILABLE = True
except ImportError:
    _email_svc = None
    _EMAIL_AVAILABLE = False


def _send_email(fn_name: str, *args, **kwargs) -> bool:
    """
    Fire-and-forget wrapper around core.email_service.<fn_name>.
    Never raises — a broken email must never abort a transaction view.
    """
    if not _EMAIL_AVAILABLE:
        return False
    try:
        return getattr(_email_svc, fn_name)(*args, **kwargs)
    except Exception as _exc:
        import logging
        logging.getLogger("transactions.email").warning(
            "_send_email(%s) failed: %s", fn_name, _exc
        )
        return False


def _member_emails_on(library) -> bool:
    """
    Master toggle: True if transactional member emails are enabled.
    Reads NotificationSettings.email_overdue_reminder as the shared
    'send member emails' flag (add a dedicated field if you need finer control).
    Also requires the subscription to be active.
    """
    if not _subscription_is_active(library):
        return False
    try:
        return bool(library.notifications.email_overdue_reminder)
    except Exception:
        return False


def _member_sms_on(library) -> bool:
    """
    True if SMS reminders are enabled in both NotificationSettings
    AND the library's current plan includes allow_sms.
    """
    if not _subscription_is_active(library):
        return False
    if not _plan_allows(library, "sms"):
        return False
    try:
        return bool(library.notifications.sms_reminder)
    except Exception:
        return False


# ─────────────────────────────────────────────────────────────────────────────
# Internal helpers
# ─────────────────────────────────────────────────────────────────────────────

def _get_library_or_404(request):
    """Return the Library for the logged-in user or raise Http404."""
    try:
        return request.user.library
    except Exception:
        raise Http404("No library associated with this account.")


def _get_library_rules(library):
    """Safely return library.rules or None."""
    try:
        return library.rules
    except Exception:
        return None


def _get_borrow_limit(library, member=None) -> int:
    """
    Role-specific borrow limit from LibraryRuleSettings.
    Returns 0 when no limit is configured (= unlimited).
    """
    role = getattr(member, "role", "") if member else ""
    try:
        rules = library.rules
        if role in ("teacher", "faculty", "staff"):
            limit = getattr(rules, "teacher_borrow_limit", None)
        else:
            limit = getattr(rules, "student_borrow_limit", None)
        if limit is None:
            limit = getattr(rules, "max_books_per_member", None)
        if limit is not None:
            return int(limit)
    except Exception:
        pass
    try:
        from accounts.models import MemberSettings
        ms = MemberSettings.objects.get(library=library)
        if ms.borrow_limit:
            return int(ms.borrow_limit)
    except Exception:
        pass
    return 0


def _get_max_renewals(library) -> int:
    """Max renewals from library rules; falls back to 2."""
    try:
        rules = library.rules
        for attr in ("max_renewal_count", "max_renewals", "renewal_limit"):
            val = getattr(rules, attr, None)
            if val is not None:
                return int(val)
    except Exception:
        pass
    return 2


def _get_fine_rate(library, txn=None) -> Decimal:
    """Live fine rate from library rules; falls back to txn snapshot then ₹2."""
    try:
        rate = library.rules.late_fine
        if rate is not None:
            return Decimal(rate)
    except Exception:
        pass
    if txn is not None:
        return txn.fine_rate_per_day
    return Decimal("2.00")


def _auto_fine_enabled(library) -> bool:
    """Return True if the library has auto-fine enabled (default True)."""
    try:
        return bool(library.rules.auto_fine)
    except Exception:
        return True


def _renewal_allowed(library) -> bool:
    """Return True if the library has renewals enabled (default True)."""
    try:
        return bool(library.rules.allow_renewal)
    except Exception:
        return True


def _partial_payment_allowed(library) -> bool:
    """Return True if partial fine payments are allowed (default False)."""
    try:
        return bool(library.rules.allow_partial_payment)
    except Exception:
        return False


def _auto_mark_lost_enabled(library) -> bool:
    """Return True if the library auto-marks severely overdue books as lost."""
    try:
        return bool(library.rules.auto_mark_lost)
    except Exception:
        return False


def _advance_booking_allowed(library) -> bool:
    """Return True if the library allows advance booking (default False)."""
    try:
        return bool(library.rules.allow_advance_booking)
    except Exception:
        return False


def _get_grace_period(library) -> int:
    """Return the grace period in days before a fine begins to accrue (default 0)."""
    try:
        return int(library.rules.grace_period or 0)
    except Exception:
        return 0


# ─────────────────────────────────────────────────────────────────────────────
# Subscription / Plan enforcement
# ─────────────────────────────────────────────────────────────────────────────
#
# _get_subscription(library)
#   Returns the subscriptions.Subscription for the library, or None.
#
# _subscription_is_active(library)
#   True only when status == 'active' AND expiry_date >= today.
#   If no subscription record exists at all → treated as inactive.
#
# _plan_allows(library, feature)
#   Checks a boolean feature flag on the linked Plan object.
#   Falls back to False when no subscription / plan is found.
#
# _require_active_subscription(request, library)
#   Call at the top of any view that requires an active subscription.
#   Returns None when OK, or an HttpResponse redirect with an error
#   message when the subscription is inactive / expired.
#
# _require_plan_feature(request, library, feature, feature_label)
#   Same pattern but checks a specific plan feature flag.
#   Returns None when OK, or an HttpResponse redirect when the feature
#   is not included in the library's current plan.
#
# Feature flag → Plan field mapping
# ─────────────────────────────────
#   "reports"          → Plan.allow_reports
#   "export"           → Plan.allow_export
#   "sms"              → Plan.allow_sms
#   "email_reminders"  → Plan.allow_email_reminders
#   "api_access"       → Plan.allow_api_access
#   "custom_branding"  → Plan.allow_custom_branding
#   "advance_booking"  → Plan.allow_advance_booking
#   "priority_support" → Plan.priority_support
# ─────────────────────────────────────────────────────────────────────────────

_PLAN_FEATURE_MAP = {
    "reports":          "allow_reports",
    "export":           "allow_export",
    "sms":              "allow_sms",
    "email_reminders":  "allow_email_reminders",
    "api_access":       "allow_api_access",
    "custom_branding":  "allow_custom_branding",
    "advance_booking":  "allow_advance_booking",
    "priority_support": "priority_support",
}


def _get_subscription(library):
    """Return the subscriptions.Subscription for this library, or None."""
    try:
        return library.subscription_detail
    except Exception:
        return None


def _subscription_is_active(library) -> bool:
    """
    True only when a Subscription exists, status == 'active',
    and expiry_date >= today.
    """
    sub = _get_subscription(library)
    if sub is None:
        return False
    return sub.is_active   # property: status == 'active' AND expiry >= today


def _plan_allows(library, feature: str) -> bool:
    """
    Check whether the library's current plan includes *feature*.
    feature must be a key in _PLAN_FEATURE_MAP.
    Returns False when no subscription / plan exists.
    """
    sub = _get_subscription(library)
    if sub is None:
        return False
    plan = getattr(sub, "plan", None)
    if plan is None:
        return False
    field = _PLAN_FEATURE_MAP.get(feature)
    if field is None:
        return False
    return bool(getattr(plan, field, False))


def _require_active_subscription(request, library):
    """
    Returns None if the subscription is active.
    Returns an HttpResponse (redirect) with an error message if not.
    """
    if _subscription_is_active(library):
        return None

    sub = _get_subscription(library)
    if sub is None:
        msg = (
            "Your library does not have an active subscription. "
            "Please choose a plan to continue."
        )
    elif sub.expiry_date < date.today():
        msg = (
            f"Your subscription expired on {sub.expiry_date.strftime('%d %b %Y')}. "
            "Please renew your plan to continue using this feature."
        )
    else:
        msg = (
            f"Your subscription is currently {sub.get_status_display().lower()}. "
            "Please contact the platform administrator or renew your plan."
        )

    messages.error(request, msg)
    return redirect("accounts:settings")


def _require_plan_feature(request, library, feature: str, feature_label: str):
    """
    Returns None if the plan includes *feature*.
    Returns an HttpResponse (redirect) with an upgrade message if not.
    Implicitly also requires an active subscription.
    """
    # Active check first
    guard = _require_active_subscription(request, library)
    if guard is not None:
        return guard

    if _plan_allows(library, feature):
        return None

    sub  = _get_subscription(library)
    plan = getattr(sub, "plan", None)
    plan_name = plan.name if plan else "your current plan"

    messages.error(
        request,
        f"'{feature_label}' is not available on the {plan_name} plan. "
        "Please upgrade your subscription to access this feature.",
    )
    return redirect("subscriptions:plans")


def _member_has_overdue_loan(member, library) -> bool:
    """
    Rule 5: Return True if the member has ANY active overdue transaction.
    An overdue member may not issue or renew books, and must return all loans.
    """
    return Transaction.objects.for_library(library).filter(
        member=member,
        status=Transaction.STATUS_OVERDUE,
    ).exists()


def _member_is_blocked(member) -> bool:
    """
    Rule 6: Return True if the member account is blocked.
    Blocked status is stored on Member.status == 'blocked'.
    """
    return getattr(member, "status", "active") == "blocked"


def _block_member_if_overdue(member, library) -> bool:
    """
    Rule 6: Block the member if they have overdue books that have not been
    returned. Called after sync. Returns True if the member was blocked.
    """
    if _member_has_overdue_loan(member, library):
        if getattr(member, "status", "active") not in ("blocked", "inactive", "passout"):
            try:
                member.status = "blocked"
                member.save(update_fields=["status"])
                # ── Email: account blocked notification ────────────────────────
                if _member_emails_on(library) and getattr(member, "email", None):
                    _send_email("send_member_blocked_email", member)
                return True
            except Exception:
                pass
    return False


def _has_unpaid_fine(txn, library) -> tuple[bool, Decimal]:
    """
    Query Fine table directly (authoritative source).
    Returns (has_unpaid: bool, total_unpaid: Decimal).
    """
    qs = Fine.objects.for_library(library).filter(
        transaction=txn, status=Fine.STATUS_UNPAID
    )
    if not qs.exists():
        return False, Decimal("0.00")
    total = qs.aggregate(t=Sum("amount"))["t"] or Decimal("0.00")
    return True, total


def _create_overdue_fine(txn, library, overdue_days: int, fine_rate: Decimal) -> Fine | None:
    """
    Create a TYPE_OVERDUE Fine for the given number of overdue days.
    Returns the Fine instance, or None if overdue_days <= 0.
    """
    if overdue_days <= 0:
        return None
    amount = Decimal(overdue_days) * fine_rate
    return Fine.objects.create(
        library     = library,
        transaction = txn,
        fine_type   = Fine.TYPE_OVERDUE,
        amount      = amount,
        status      = Fine.STATUS_UNPAID,
    )


def _wants_json(request) -> bool:
    return (
        request.GET.get("format") == "json"
        or "application/json" in request.headers.get("Accept", "")
    )


# ── Throttled overdue sync ─────────────────────────────────────────────────────
_last_sync: dict[int, float] = {}
STALE_THRESHOLD_SECONDS: int = 90


def _sync_overdue_if_stale(library) -> None:
    now  = time.monotonic()
    last = _last_sync.get(library.pk, 0.0)
    if now - last >= STALE_THRESHOLD_SECONDS:
        Transaction.sync_overdue_for_library(library)
        _last_sync[library.pk] = now
        # Bulk-settle overdue transactions whose fines are all paid
        _sync_overdue_settled_for_library(library)
        # Rule: auto-mark severely overdue books as lost (if toggle ON)
        _auto_mark_lost_overdue_books(library)
        # Rule 6: block members who have overdue loans that are not returned
        _auto_block_overdue_members(library)


def _auto_mark_lost_overdue_books(library) -> None:
    """
    Rule: if auto_mark_lost is ON, automatically mark transactions as STATUS_LOST
    when a book is severely overdue (overdue_days > 30 * max_renewal_count, or
    a flat 60-day threshold). Called from the sync path.
    """
    if not _auto_mark_lost_enabled(library):
        return
    try:
        AUTO_LOST_THRESHOLD_DAYS = 60
        cutoff = date.today() - timedelta(days=AUTO_LOST_THRESHOLD_DAYS)
        overdue_qs = Transaction.objects.for_library(library).filter(
            status=Transaction.STATUS_OVERDUE,
            due_date__lt=cutoff,
        ).select_related("book")
        for txn in overdue_qs:
            try:
                from django.db import transaction as _dbt
                with _dbt.atomic():
                    txn.status    = Transaction.STATUS_LOST
                    txn.lost_date = date.today()
                    txn.notes     = (txn.notes or "") + " [Auto-marked lost: overdue > 60 days]"
                    txn.save(update_fields=["status", "lost_date", "notes", "updated_at"])
                    book = txn.book
                    book.total_copies     = max(0, book.total_copies - 1)
                    book.available_copies = min(book.available_copies, book.total_copies)
                    book.save(update_fields=["total_copies", "available_copies"])
                    from .models import MissingBook
                    MissingBook.objects.update_or_create(
                        transaction=txn,
                        defaults={
                            "library":       library,
                            "book":          book,
                            "status":        MissingBook.STATUS_LOST,
                            "reported_date": date.today(),
                            "notes":         "Auto-marked lost: overdue > 60 days",
                        },
                    )
                    if _auto_fine_enabled(library):
                        from finance.models import Fine as _Fine
                        _Fine.objects.get_or_create(
                            library=library,
                            transaction=txn,
                            fine_type=_Fine.TYPE_LOST,
                            defaults={"amount": txn.fine_amount or Decimal("0.00"), "status": _Fine.STATUS_UNPAID},
                        )
                # ── Email: book auto-marked lost ───────────────────────────────
                if _member_emails_on(library):
                    try:
                        _member = txn.member if hasattr(txn, "member") else None
                        if _member and getattr(_member, "email", None):
                            _send_email("send_book_lost_email", _member, txn)
                    except Exception:
                        pass
            except Exception as exc:
                import logging
                logging.getLogger("transactions.views").warning(
                    "auto_mark_lost: failed for txn %s: %s", txn.pk, exc
                )
    except Exception:
        pass


def _auto_block_overdue_members(library) -> None:
    """
    Rule 6: For every member with at least one STATUS_OVERDUE transaction,
    flip their status to 'blocked' so issue / renew are prevented until
    they return all overdue books.
    Only touches members whose current status is 'active'.
    """
    from members.models import Member
    try:
        overdue_member_ids = set(
            Transaction.objects.for_library(library)
            .filter(status=Transaction.STATUS_OVERDUE)
            .values_list("member_id", flat=True)
            .distinct()
        )
        if overdue_member_ids:
            # Collect members about to be blocked so we can email them
            if _member_emails_on(library):
                members_to_notify = list(
                    Member.objects.filter(
                        pk__in=overdue_member_ids,
                        status="active",
                    ).exclude(email="").only("pk", "email", "first_name", "last_name", "member_id", "status")
                )
            else:
                members_to_notify = []

            Member.objects.filter(
                pk__in=overdue_member_ids,
                status="active",
            ).update(status="blocked")

            # Send blocked notification to each newly blocked member
            for _m in members_to_notify:
                _m.status = "blocked"  # reflect new state for email content
                _send_email("send_member_blocked_email", _m)
    except Exception:
        pass


def _sync_overdue_settled_for_library(library) -> None:
    """
    Bulk-flip overdue → overdue_settled for transactions where:
      1. At least one Fine row exists (paid or waived), AND
      2. No UNPAID Fine rows remain.

    This prevents transactions with zero fines (fine not yet created)
    from being incorrectly marked as settled.
    """
    today = date.today()

    # Collect all transaction PKs that are currently OVERDUE for this library
    overdue_pks = set(
        Transaction.objects.for_library(library)
        .filter(status=Transaction.STATUS_OVERDUE)
        .values_list("pk", flat=True)
    )
    if not overdue_pks:
        return

    # Transactions that still have at least one UNPAID fine → must NOT settle
    has_unpaid = set(
        Fine.objects.for_library(library)
        .filter(transaction_id__in=overdue_pks, status=Fine.STATUS_UNPAID)
        .values_list("transaction_id", flat=True)
        .distinct()
    )

    # Transactions that have at least one PAID or WAIVED fine → eligible
    has_paid = set(
        Fine.objects.for_library(library)
        .filter(
            transaction_id__in=overdue_pks,
            status__in=(Fine.STATUS_PAID, Fine.STATUS_WAIVED),
        )
        .values_list("transaction_id", flat=True)
        .distinct()
    )

    # Settle only if: has a paid/waived fine AND has no unpaid fine
    settled_pks = has_paid - has_unpaid
    if not settled_pks:
        return

    Transaction.objects.filter(pk__in=settled_pks).update(
        status         = Transaction.STATUS_OVERDUE_SETTLED,
        fine_paid      = True,
        fine_paid_date = today,
    )


def _reset_overdue_after_fine_settled(txn, library) -> None:
    """
    After a Fine is paid or waived: if the transaction is still OVERDUE,
    has no remaining UNPAID fines, AND has at least one PAID/WAIVED fine,
    flip it to OVERDUE_SETTLED immediately (instant feedback at payment time).
    """
    if txn.status != Transaction.STATUS_OVERDUE:
        return
    has_unpaid, _ = _has_unpaid_fine(txn, library)
    if has_unpaid:
        return
    # Guard: only settle if at least one paid/waived fine exists
    has_paid = Fine.objects.for_library(library).filter(
        transaction=txn,
        status__in=(Fine.STATUS_PAID, Fine.STATUS_WAIVED),
    ).exists()
    if not has_paid:
        return
    today = date.today()
    txn.status         = Transaction.STATUS_OVERDUE_SETTLED
    txn.fine_paid      = True
    txn.fine_paid_date = today
    txn.save(update_fields=["status", "fine_paid", "fine_paid_date", "updated_at"])


# ─────────────────────────────────────────────────────────────────────────────
# Helper: upsert a Fine row without creating duplicates
# ─────────────────────────────────────────────────────────────────────────────

def _upsert_fine(
    library,
    txn,
    fine_type: str,
    amount: Decimal,
    status: str,
    paid_date=None,
) -> Fine:
    """
    Create-or-update a Fine row keyed on (library, transaction, fine_type).

    Rules:
      • If no row exists → create it with the supplied values.
      • If an UNPAID row exists → update amount / status / paid_date in-place.
      • If a PAID or WAIVED row already exists → leave it untouched and return it.

    This is the single canonical path used by both return_book (view) and
    fine_sync (background thread) so neither can produce a duplicate row.
    """
    fine_obj, created = Fine.objects.get_or_create(
        library     = library,
        transaction = txn,
        fine_type   = fine_type,
        defaults={
            "amount":    amount,
            "status":    status,
            "paid_date": paid_date,
        },
    )
    if not created and fine_obj.status == Fine.STATUS_UNPAID:
        changed = False
        if fine_obj.amount != amount:
            fine_obj.amount = amount
            changed = True
        if fine_obj.status != status:
            fine_obj.status = status
            changed = True
        if fine_obj.paid_date != paid_date:
            fine_obj.paid_date = paid_date
            changed = True
        if changed:
            fine_obj.save(update_fields=["amount", "status", "paid_date", "updated_at"])
    return fine_obj


# ─────────────────────────────────────────────────────────────────────────────
# 1. Transaction List
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def transaction_list(request):
    library = _get_library_or_404(request)
    guard = _require_active_subscription(request, library)
    if guard: return guard
    _sync_overdue_if_stale(library)

    from django.db.models import OuterRef, Subquery, Sum as _Sum, DecimalField as _DCF
    from finance.models import Fine as _Fine

    qs = Transaction.objects.for_library(library).select_related("member", "book")

    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(
            Q(member__first_name__icontains=q)
            | Q(member__last_name__icontains=q)
            | Q(book__title__icontains=q)
            | Q(pk__icontains=q)
        )

    status = request.GET.get("status", "").strip()
    if status:
        qs = qs.filter(status=status)

    date_from = request.GET.get("date_from", "")
    date_to   = request.GET.get("date_to",   "")
    if date_from:
        qs = qs.filter(issue_date__gte=date_from)
    if date_to:
        qs = qs.filter(issue_date__lte=date_to)

    # Annotate each transaction with the real total fine from the Fine table.
    # This replaces txn.fine_amount (a live property = overdue_days × rate)
    # which returns 0 for returned/settled books even when fines were charged.
    fine_subq = (
        _Fine.objects
        .filter(transaction=OuterRef("pk"))
        .values("transaction")
        .annotate(total=_Sum("amount", output_field=_DCF()))
        .values("total")
    )
    qs = qs.annotate(db_total_fine=Subquery(fine_subq, output_field=_DCF()))

    base           = Transaction.objects.for_library(library)
    total_count    = base.count()
    issued_count   = base.filter(status=Transaction.STATUS_ISSUED).count()
    overdue_count  = base.filter(status=Transaction.STATUS_OVERDUE).count()
    returned_count = base.filter(status=Transaction.STATUS_RETURNED).count()
    lost_count     = base.filter(status=Transaction.STATUS_LOST).count()

    per_page     = int(request.GET.get("per_page", 5))
    paginator    = Paginator(qs.order_by("-created_at"), per_page)
    transactions = paginator.get_page(request.GET.get("page", 1))

    return render(request, "transactions/transaction_list.html", {
        "transactions":   transactions,
        "total_count":    total_count,
        "issued_count":   issued_count,
        "overdue_count":  overdue_count,
        "returned_count": returned_count,
        "lost_count":     lost_count,
    })


# ─────────────────────────────────────────────────────────────────────────────
# 2. Transaction Detail
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def transaction_detail(request, pk):
    library = _get_library_or_404(request)
    guard = _require_active_subscription(request, library)
    if guard: return guard
    txn = get_object_or_404(
        Transaction.objects.for_library(library).select_related("member", "book"),
        pk=pk,
    )
    fines = list(txn.fines.all().order_by("-created_at"))

    member             = txn.member
    active_loans_count = Transaction.objects.for_library(library).filter(
        member=member,
        status__in=(Transaction.STATUS_ISSUED, Transaction.STATUS_OVERDUE),
    ).count()

    # ── Fine breakdown from finance_fine table (single query set, reused) ──
    fine_qs = Fine.objects.for_library(library).filter(transaction=txn)

    def _sum(qs):
        return qs.aggregate(t=Sum("amount", output_field=_DF()))["t"] or Decimal("0.00")

    # Per-type totals — used for the breakdown rows in the Fine Details card
    db_overdue_fine  = _sum(fine_qs.filter(fine_type=Fine.TYPE_OVERDUE))
    db_damage_charge = _sum(fine_qs.filter(fine_type=Fine.TYPE_DAMAGE))
    db_lost_penalty  = _sum(fine_qs.filter(fine_type=Fine.TYPE_LOST))

    # Per-status totals — used for Amount Due / Already Paid rows
    unpaid_fine_amount = _sum(fine_qs.filter(status=Fine.STATUS_UNPAID))
    paid_fine_amount   = _sum(fine_qs.filter(status=Fine.STATUS_PAID))

    # Outstanding fine = all unpaid fines across ALL of this member's transactions
    # (shown in the Member info card — may differ from unpaid_fine_amount above)
    outstanding_fine = (
        Fine.objects.for_library(library)
        .filter(transaction__member=member, status=Fine.STATUS_UNPAID)
        .aggregate(t=Sum("amount", output_field=_DF()))["t"] or Decimal("0.00")
    )

    return render(request, "transactions/transaction_detail.html", {
        "transaction":        txn,
        "fines":              fines,
        "active_loans_count": active_loans_count,
        "borrow_limit":       _get_borrow_limit(library, member),
        "outstanding_fine":   outstanding_fine,
        "max_renewals":       _get_max_renewals(library),
        # ── Fine Details card context ──────────────────────────────────────
        "db_overdue_fine":    db_overdue_fine,
        "db_damage_charge":   db_damage_charge,
        "db_lost_penalty":    db_lost_penalty,
        "unpaid_fine_amount": unpaid_fine_amount,
        "paid_fine_amount":   paid_fine_amount,
    })


# ─────────────────────────────────────────────────────────────────────────────
# 3. Issue Book
# ─────────────────────────────────────────────────────────────────────────────
#
# RULES ENFORCED:
#   ✔ Member active (not blocked / inactive / passout)
#   ✔ Member has no overdue loans (Rule 5)
#   ✔ Book available (copy status + available_copies)
#   ✔ Borrow limit not exceeded
#   ✔ No unpaid fines (Rule 1)
#   ✔ Auto-fine setting respected when creating fines (Rule 7)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def issue_book(request):
    library = _get_library_or_404(request)
    guard = _require_active_subscription(request, library)
    if guard: return guard
    owner   = library.user
    rules   = _get_library_rules(library)

    from members.models import Member
    from books.models import Book

    if not rules:
        messages.error(request, "Library rule settings not configured.")
        return redirect("transactions:transaction_list")

    borrowing_period  = int(rules.borrowing_period or 14)
    fine_rate_per_day = rules.late_fine or Decimal("0.00")
    max_renewal_count = _get_max_renewals(library)
    max_books_per_member = _get_borrow_limit(library)

    student_borrow_limit = getattr(rules, "student_borrow_limit", None)
    teacher_borrow_limit = getattr(rules, "teacher_borrow_limit", None)
    is_institute         = bool(student_borrow_limit or teacher_borrow_limit)
    student_borrow_limit = student_borrow_limit or max_books_per_member
    teacher_borrow_limit = teacher_borrow_limit or max_books_per_member

    today            = date.today()
    default_due_date = today + timedelta(days=borrowing_period)

    # ── Member list annotated with active loan count ───────────────────────────
    _members_qs = (
        Member.objects
        .filter(owner=owner, status="active")
        .annotate(active_loans_count=Count(
            "issue_transactions",
            filter=Q(
                issue_transactions__library=library,
                issue_transactions__status__in=(
                    Transaction.STATUS_ISSUED,
                    Transaction.STATUS_OVERDUE,
                ),
            ),
        ))
        .order_by("first_name", "last_name")
    )

    # Bulk unpaid fine totals — no N+1
    _fine_totals = {
        row["transaction__member_id"]: row["total"]
        for row in Fine.objects.for_library(library)
        .filter(status=Fine.STATUS_UNPAID)
        .values("transaction__member_id")
        .annotate(total=Sum("amount", output_field=_DF()))
    }

    members = []
    for m in _members_qs:
        _limit            = _get_borrow_limit(library, m)
        m.slots_available = max(0, _limit - m.active_loans_count) if _limit else -1
        m.total_due       = _fine_totals.get(m.pk, Decimal("0.00"))
        members.append(m)

    books = Book.objects.filter(owner=owner, available_copies__gt=0).order_by("title")

    # ── POST ──────────────────────────────────────────────────────────────────
    if request.method == "POST":
        form = IssueBookForm(request.POST, library=library)

        if form.is_valid():
            cd         = form.cleaned_data
            member     = cd["member"]
            book       = cd["book"]
            book_copy  = cd.get("book_copy")
            issue_date = cd["issue_date"]
            due_date   = issue_date + timedelta(days=borrowing_period)

            # ── Rule: advance booking (future issue_date) requires toggle ──────
            if cd["issue_date"] > date.today() and not _advance_booking_allowed(library):
                messages.error(
                    request,
                    "Cannot issue — advance booking (future issue date) is disabled. "
                    "Enable it in Settings → Fine & Loans → Allow Advance Booking.",
                )
                return render(request, "transactions/issue_book.html", {
                    "form": form, "members": members, "books": books,
                    "today": today.isoformat(), "default_due_date": default_due_date,
                    "rules": rules, "fine_rate_per_day": fine_rate_per_day,
                    "default_loan_days": borrowing_period,
                    "max_books_per_member": max_books_per_member,
                    "max_renewal_count": max_renewal_count,
                    "is_institute": is_institute,
                    "student_borrow_limit": student_borrow_limit,
                    "teacher_borrow_limit": teacher_borrow_limit,
                })

            # ── Rule: member active and not blocked ────────────────────────────
            if _member_is_blocked(member):
                messages.error(
                    request,
                    f"Cannot issue — {member.first_name} {member.last_name}'s account "
                    f"is blocked. They must return all overdue books first.",
                )
                return render(request, "transactions/issue_book.html", {
                    "form": form, "members": members, "books": books,
                    "today": today.isoformat(), "default_due_date": default_due_date,
                    "rules": rules, "fine_rate_per_day": fine_rate_per_day,
                    "default_loan_days": borrowing_period,
                    "max_books_per_member": max_books_per_member,
                    "max_renewal_count": max_renewal_count,
                    "is_institute": is_institute,
                    "student_borrow_limit": student_borrow_limit,
                    "teacher_borrow_limit": teacher_borrow_limit,
                })

            if getattr(member, "status", "active") != "active":
                messages.error(
                    request,
                    f"{member.first_name} {member.last_name} is not an active member "
                    f"(status: {member.get_status_display()}).",
                )
                return render(request, "transactions/issue_book.html", {
                    "form": form, "members": members, "books": books,
                    "today": today.isoformat(), "default_due_date": default_due_date,
                    "rules": rules, "fine_rate_per_day": fine_rate_per_day,
                    "default_loan_days": borrowing_period,
                    "max_books_per_member": max_books_per_member,
                    "max_renewal_count": max_renewal_count,
                    "is_institute": is_institute,
                    "student_borrow_limit": student_borrow_limit,
                    "teacher_borrow_limit": teacher_borrow_limit,
                })

            # ── Rule 5: member must not have any overdue loans ────────────────
            if _member_has_overdue_loan(member, library):
                messages.error(
                    request,
                    f"Cannot issue — {member.first_name} {member.last_name} has "
                    f"overdue books. They must return all overdue loans before "
                    f"borrowing new books.",
                )
                return render(request, "transactions/issue_book.html", {
                    "form": form, "members": members, "books": books,
                    "today": today.isoformat(), "default_due_date": default_due_date,
                    "rules": rules, "fine_rate_per_day": fine_rate_per_day,
                    "default_loan_days": borrowing_period,
                    "max_books_per_member": max_books_per_member,
                    "max_renewal_count": max_renewal_count,
                    "is_institute": is_institute,
                    "student_borrow_limit": student_borrow_limit,
                    "teacher_borrow_limit": teacher_borrow_limit,
                })

            # ── Rule: no pending fine ──────────────────────────────────────────
            has_fine, fine_total = _has_unpaid_fine_for_member(member, library)
            if has_fine:
                messages.error(
                    request,
                    f"Cannot issue — {member.first_name} {member.last_name} "
                    f"has an unpaid fine of ₹{fine_total}. Please clear it first.",
                )
                return render(request, "transactions/issue_book.html", {
                    "form": form, "members": members, "books": books,
                    "today": today.isoformat(), "default_due_date": default_due_date,
                    "rules": rules, "fine_rate_per_day": fine_rate_per_day,
                    "default_loan_days": borrowing_period,
                    "max_books_per_member": max_books_per_member,
                    "max_renewal_count": max_renewal_count,
                    "is_institute": is_institute,
                    "student_borrow_limit": student_borrow_limit,
                    "teacher_borrow_limit": teacher_borrow_limit,
                })

            # ── Rule: borrow limit ─────────────────────────────────────────────
            borrow_limit = _get_borrow_limit(library, member)
            if borrow_limit:
                active_loans = Transaction.objects.for_library(library).filter(
                    member=member,
                    status__in=(Transaction.STATUS_ISSUED, Transaction.STATUS_OVERDUE),
                ).count()
                if active_loans >= borrow_limit:
                    messages.error(
                        request,
                        f"Cannot issue — {member.first_name} {member.last_name} "
                        f"has reached their borrow limit ({borrow_limit} books).",
                    )
                    return render(request, "transactions/issue_book.html", {
                        "form": form, "members": members, "books": books,
                        "today": today.isoformat(), "default_due_date": default_due_date,
                        "rules": rules, "fine_rate_per_day": fine_rate_per_day,
                        "default_loan_days": borrowing_period,
                        "max_books_per_member": max_books_per_member,
                        "max_renewal_count": max_renewal_count,
                        "is_institute": is_institute,
                        "student_borrow_limit": student_borrow_limit,
                        "teacher_borrow_limit": teacher_borrow_limit,
                    })

            with db_transaction.atomic():
                txn = Transaction.objects.create(
                    library            = library,
                    member             = member,
                    book               = book,
                    book_copy          = book_copy,
                    issue_date         = issue_date,
                    due_date           = due_date,
                    loan_duration_days = borrowing_period,
                    fine_rate_per_day  = fine_rate_per_day,
                    status             = Transaction.STATUS_ISSUED,
                    issued_by          = request.user.get_full_name() or request.user.username,
                    notes              = cd.get("notes", ""),
                )

                # Mark copy as issued
                if book_copy is not None:
                    book_copy.status = "issued"
                    book_copy.save(update_fields=["status"])
                else:
                    if hasattr(book, "available_copies"):
                        book.available_copies = max(0, book.available_copies - 1)
                        book.save(update_fields=["available_copies"])

            messages.success(
                request,
                f'"{book.title}" issued to {member.first_name} {member.last_name}. '
                f'Due: {due_date.strftime("%d %B %Y")}.',
            )
            # ── Email: book issued confirmation ────────────────────────────────
            if _member_emails_on(library) and getattr(member, "email", None):
                _send_email("send_book_issued_email", member, txn)

            return redirect("transactions:transaction_detail", pk=txn.pk)

        else:
            for field, errs in form.errors.items():
                for e in errs:
                    messages.error(request, e)

    else:
        form = IssueBookForm(library=library)

    return render(request, "transactions/issue_book.html", {
        "form":                 form,
        "members":              members,
        "books":                books,
        "today":                today.isoformat(),
        "default_due_date":     default_due_date,
        "rules":                rules,
        "fine_rate_per_day":    fine_rate_per_day,
        "default_loan_days":    borrowing_period,
        "max_books_per_member": max_books_per_member,
        "max_renewal_count":    max_renewal_count,
        "is_institute":         is_institute,
        "student_borrow_limit": student_borrow_limit,
        "teacher_borrow_limit": teacher_borrow_limit,
    })


def _has_unpaid_fine_for_member(member, library) -> tuple[bool, Decimal]:
    """Any unpaid fine across ALL transactions for this member."""
    qs = (
        Fine.objects.for_library(library)
        .filter(transaction__member=member, status=Fine.STATUS_UNPAID)
    )
    if not qs.exists():
        return False, Decimal("0.00")
    total = qs.aggregate(t=Sum("amount"))["t"] or Decimal("0.00")
    return True, total


# ─────────────────────────────────────────────────────────────────────────────
# 4. Return Book
# ─────────────────────────────────────────────────────────────────────────────
#
# RULES ENFORCED:
#   ✔ Transaction must belong to this library
#   ✔ Not already returned
#   ✔ Late return  → Fine(TYPE_OVERDUE) upserted via _upsert_fine (Rule 7)
#   ✔ Damaged      → Fine(TYPE_DAMAGE)  upserted via _upsert_fine (Rule 7)
#   ✔ On successful return — unblock member if no remaining overdue loans (Rule 6)
#
# FIX: replaced Fine.objects.create() with _upsert_fine() so that if
# fine_sync already created a row for this (library, transaction, fine_type)
# the view updates it in-place instead of inserting a duplicate.
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def return_book(request, pk):
    library = _get_library_or_404(request)
    guard = _require_active_subscription(request, library)
    if guard: return guard
    txn = get_object_or_404(
        Transaction.objects.for_library(library).select_related("member", "book"),
        pk=pk,
    )

    if txn.status == Transaction.STATUS_RETURNED:
        messages.info(request, "This book has already been returned.")
        return redirect("transactions:transaction_detail", pk=txn.pk)

    today = date.today()

    if request.method == "POST":
        form = ReturnBookForm(request.POST)
        if form.is_valid():
            cd            = form.cleaned_data
            return_date   = cd.get("return_date") or today
            condition     = cd.get("condition", "good")
            damage_charge = cd.get("damage_charge") or Decimal("0.00")
            return_notes  = cd.get("return_notes", "")
            fine_paid_now = cd.get("fine_paid_now", False)

            # For damaged/lost the form always sends fine_paid_now=1
            # so payment is captured at the same moment as the return.
            is_damaged = condition == Transaction.CONDITION_DAMAGED
            is_lost    = condition == "lost"   # not a model constant but valid condition value

            # ── Rule: late return → compute overdue fine (with grace period) ──
            grace_period = _get_grace_period(library)
            overdue_days = max(0, (return_date - txn.due_date).days - grace_period)
            fine_rate    = _get_fine_rate(library, txn)
            overdue_fine = Decimal(overdue_days) * fine_rate if overdue_days > 0 else Decimal("0.00")

            fine_status = Fine.STATUS_PAID if fine_paid_now else Fine.STATUS_UNPAID
            fine_date   = return_date if fine_paid_now else None

            # ── Rule 7: only create fines if auto_fine is enabled ──────────────
            apply_fines = _auto_fine_enabled(library)

            with db_transaction.atomic():
                # ── Lost returned inline — treat as lost + return in one shot ──
                if is_lost:
                    txn.status    = Transaction.STATUS_LOST
                    txn.lost_date = return_date
                else:
                    txn.status = Transaction.STATUS_RETURNED

                txn.return_date      = return_date
                txn.return_condition = condition
                txn.damage_charge    = damage_charge
                txn.return_notes     = return_notes
                txn.returned_to      = request.user.get_full_name() or request.user.username
                # NOTE: fine_paid / fine_paid_date are set AFTER payment
                # is collected on the finance:process_payment screen.
                txn.save()

                # ── Restore / update copy status based on condition ────────────
                book       = txn.book
                from books.models import BookCopy as _BookCopy
                try:
                    copy_to_return = txn.book_copy or (
                        _BookCopy.objects
                        .filter(book=book, status__in=(_BookCopy.Status.BORROWED, "issued", "borrowed"))
                        .order_by("updated_at")
                        .first()
                    )
                    if copy_to_return:
                        if is_damaged:
                            copy_to_return.status      = _BookCopy.Status.DAMAGED
                            copy_to_return.returned_at = return_date
                        elif is_lost:
                            copy_to_return.status = _BookCopy.Status.LOST
                        else:
                            copy_to_return.status      = _BookCopy.Status.AVAILABLE
                            copy_to_return.returned_at = return_date
                        copy_to_return.save(update_fields=["status", "returned_at", "updated_at"])
                except Exception:
                    pass

                if hasattr(book, "available_copies") and hasattr(book, "total_copies"):
                    update_fields = ["total_copies", "available_copies"]
                    if is_damaged or is_lost:
                        # Copy leaves inventory permanently
                        book.total_copies     = max(0, book.total_copies - 1)
                        book.available_copies = min(book.available_copies, book.total_copies)
                        # ── Sync book price if staff entered a different amount ──
                        if damage_charge and damage_charge != (book.price or Decimal("0.00")):
                            book.price = damage_charge
                            update_fields.append("price")
                    else:
                        # Normal return — restore one slot
                        book.available_copies = min(book.total_copies, book.available_copies + 1)
                    book.save(update_fields=update_fields)
                elif (is_damaged or is_lost) and damage_charge and hasattr(book, "price"):
                    # Edge case: book model has no copy-count fields but does have price
                    if damage_charge != (book.price or Decimal("0.00")):
                        book.price = damage_charge
                        book.save(update_fields=["price"])

                # ── MissingBook record for lost copies ─────────────────────────
                if is_lost:
                    MissingBook.objects.update_or_create(
                        transaction=txn,
                        defaults={
                            "library":        library,
                            "book":           book,
                            "status":         MissingBook.STATUS_LOST,
                            "reported_date":  return_date,
                            "notes":          return_notes,
                            "penalty_amount": damage_charge,
                            "penalty_reason": "lost",
                            "penalty_paid":   False,
                        },
                    )

                # ── Rule 2: fine for late return — always create as UNPAID ──────
                # Even for damaged/lost, the overdue fine goes to process_payment
                if overdue_days > 0 and apply_fines:
                    _upsert_fine(
                        library   = library,
                        txn       = txn,
                        fine_type = Fine.TYPE_OVERDUE,
                        amount    = overdue_fine,
                        status    = Fine.STATUS_UNPAID,
                        paid_date = None,
                    )

                # ── Rule 4: fine for damaged/lost — always create, even at 0 ──
                # When book price is NULL the charge arrives as 0; we still
                # create the Fine so staff can update the amount on the payment
                # screen later.
                charge_fine_obj = None
                if (is_damaged or is_lost) and apply_fines:
                    fine_type_for_charge = Fine.TYPE_LOST if is_lost else Fine.TYPE_DAMAGE
                    charge_fine_obj = _upsert_fine(
                        library   = library,
                        txn       = txn,
                        fine_type = fine_type_for_charge,
                        amount    = damage_charge,
                        status    = Fine.STATUS_UNPAID,
                        paid_date = None,
                    )

                # ── For normal (good/fair) return with fine_paid_now ───────────
                if fine_paid_now and not (is_damaged or is_lost) and apply_fines:
                    Fine.objects.for_library(library).filter(
                        transaction=txn,
                        status=Fine.STATUS_UNPAID,
                    ).update(status=Fine.STATUS_PAID, paid_date=return_date)
                    txn.fine_paid      = True
                    txn.fine_paid_date = return_date
                    txn.save(update_fields=["fine_paid", "fine_paid_date", "updated_at"])

                # ── Rule 6: unblock member if no remaining overdue loans ────────
                member = txn.member
                if not _member_has_overdue_loan(member, library):
                    if getattr(member, "status", "active") == "blocked":
                        try:
                            member.status = "active"
                            member.save(update_fields=["status"])
                        except Exception:
                            pass

            # ── Redirect: damaged/lost → process_payment for collection ────────
            if (is_damaged or is_lost) and apply_fines:
                # Find the primary charge fine to pre-load on the payment screen
                primary_fine = charge_fine_obj or (
                    Fine.objects.for_library(library)
                    .filter(transaction=txn, status=Fine.STATUS_UNPAID)
                    .order_by("-created_at")
                    .first()
                )
                if primary_fine:
                    messages.info(
                        request,
                        f'"{txn.book.title}" {"marked Lost" if is_lost else "returned as Damaged"}. '
                        f"Copy removed from inventory. Collect payment below.",
                    )
                    return redirect(
                        reverse("finance:process_payment") + f"?fine_id={primary_fine.fine_id}"
                    )

            # ── Success message (normal return or no fines) ────────────────────
            if is_lost:
                action_msg = f'"{txn.book.title}" marked as Lost. Copy removed from inventory.'
            elif is_damaged:
                action_msg = f'"{txn.book.title}" returned — copy marked Damaged and removed from inventory.'
            else:
                action_msg = f'"{txn.book.title}" returned successfully.'

            fine_msg = ""
            if overdue_days > 0 and apply_fines and not (is_damaged or is_lost):
                fine_msg += f" Overdue fine ₹{overdue_fine} {'collected' if fine_paid_now else 'recorded (unpaid)'}."

            messages.success(request, action_msg + fine_msg)

            # ── Email ──────────────────────────────────────────────────────────
            member = txn.member
            if _member_emails_on(library) and getattr(member, "email", None):
                _send_email("send_book_returned_email", member, txn, overdue_fine)

            return redirect("transactions:transaction_detail", pk=txn.pk)

    else:
        form = ReturnBookForm(initial={
            "transaction_id": txn.pk,
            "return_date":    today.isoformat(),
        })

    # Preview fine for the return form (grace period applied)
    grace_period_preview = _get_grace_period(library)
    overdue_days_preview = max(0, (today - txn.due_date).days - grace_period_preview)
    fine_rate_preview    = _get_fine_rate(library, txn)
    fine_preview         = Decimal(overdue_days_preview) * fine_rate_preview

    return render(request, "transactions/return_book.html", {
        "transaction":       txn,
        "form":              form,
        "today":             today,
        "overdue_days":      overdue_days_preview,
        "fine_preview":      fine_preview,
        "fine_rate_per_day": fine_rate_preview,
        "book_price":        txn.book.price,
    })


# ─────────────────────────────────────────────────────────────────────────────
# 5. Renew Book
# ─────────────────────────────────────────────────────────────────────────────
#
# RULES ENFORCED:
#   ✔ Member active and not blocked (Rule 6)
#   ✔ Member has no OTHER overdue loans besides this one (Rule 5)
#   ✔ Renewal count < max_renewal_count
#   ✔ No unpaid Fine rows on this transaction  (Fine table = authoritative)
#   ✔ Late renewal (renewal_date > due_date) → Fine(TYPE_OVERDUE) created
#     BEFORE extending the due_date — only if auto_fine is ON (Rule 7)
#
# NOTE: txn.fine_amount is a @property (overdue_days × rate). It must NOT
# be used as the fine check — it ignores payments. We only query Fine table.
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def renew_book(request, pk):
    if request.method != "POST":
        return redirect("transactions:transaction_detail", pk=pk)

    library   = _get_library_or_404(request)
    guard = _require_active_subscription(request, library)
    if guard: return guard
    want_json = _wants_json(request)

    txn = get_object_or_404(
        Transaction.objects.for_library(library).select_related("member", "book"),
        pk=pk,
    )

    def _error(msg, extra=None):
        if want_json:
            payload = {"ok": False, "error": msg}
            if extra:
                payload.update(extra)
            return JsonResponse(payload, status=400)
        messages.error(request, msg)
        return redirect("transactions:transaction_detail", pk=pk)

    # ── Rule: renewals must be enabled in library settings ────────────────────
    if not _renewal_allowed(library):
        return _error("Book renewals are currently disabled by the library administrator.")

    # ── Guard: only active loans ───────────────────────────────────────────────
    if txn.status not in (Transaction.STATUS_ISSUED, Transaction.STATUS_OVERDUE, Transaction.STATUS_OVERDUE_SETTLED):
        return _error(
            f"Only issued or overdue loans can be renewed. "
            f"Current status: {txn.get_status_display()}."
        )

    # ── Rule 6: member must not be blocked ────────────────────────────────────
    member = txn.member
    if _member_is_blocked(member):
        return _error(
            f"Cannot renew — {member.first_name} {member.last_name}'s account "
            f"is blocked. They must return all overdue books first."
        )

    # ── Rule: member active ────────────────────────────────────────────────────
    if getattr(member, "status", "active") != "active":
        return _error(
            f"Cannot renew — {member.first_name} {member.last_name}'s account is "
            f"not active (status: {member.get_status_display()})."
        )

    # ── Rule 5: member must not have OTHER overdue loans ──────────────────────
    other_overdue = Transaction.objects.for_library(library).filter(
        member=member,
        status=Transaction.STATUS_OVERDUE,
    ).exclude(pk=txn.pk).exists()
    if other_overdue:
        return _error(
            f"Cannot renew — {member.first_name} {member.last_name} has other "
            f"overdue books. They must return all overdue loans first."
        )

    # ── Rule: renewal limit ────────────────────────────────────────────────────
    max_renewals = _get_max_renewals(library)
    if txn.renewal_count >= max_renewals:
        return _error(
            f"Cannot renew — renewal limit of {max_renewals} already reached.",
            {"renewal_count": txn.renewal_count, "max_renewals": max_renewals},
        )

    # ── Rule: no unpaid fine (Fine table is authoritative) ────────────────────
    has_fine, fine_total = _has_unpaid_fine(txn, library)
    if has_fine:
        return _error(
            f"Cannot renew — this loan has an unpaid fine of ₹{fine_total}. "
            "Please clear the fine before renewing.",
            {"fine_amount": str(fine_total)},
        )

    # ── Rule: late renewal → create overdue Fine BEFORE extending due_date ─────
    today        = date.today()
    fine_rate    = _get_fine_rate(library, txn)
    grace_period = _get_grace_period(library)
    overdue_days = max(0, (today - txn.due_date).days - grace_period)
    new_due_date = today + timedelta(days=txn.loan_duration_days)

    with db_transaction.atomic():
        # ── Rule 7: only apply late fine if auto_fine is enabled ──────────────
        late_fine_created = None
        if overdue_days > 0 and _auto_fine_enabled(library):
            late_fine_amount  = Decimal(overdue_days) * fine_rate
            late_fine_created = Fine.objects.create(
                library     = library,
                transaction = txn,
                fine_type   = Fine.TYPE_OVERDUE,
                amount      = late_fine_amount,
                status      = Fine.STATUS_UNPAID,
            )

        # Extend the loan
        txn.renewal_count    += 1
        txn.due_date          = new_due_date
        txn.fine_rate_per_day = fine_rate        # refresh from live rule
        txn.status            = Transaction.STATUS_ISSUED   # clear overdue flag
        txn.save(update_fields=[
            "renewal_count", "due_date", "fine_rate_per_day", "status", "updated_at",
        ])

    msg = (
        f'Loan renewed. New due date: {new_due_date.strftime("%d %B %Y")} '
        f'(renewal {txn.renewal_count}/{max_renewals}).'
    )
    if late_fine_created:
        msg += (
            f" An overdue fine of ₹{late_fine_created.amount} "
            f"({overdue_days} day(s) late) has been applied."
        )

    if want_json:
        return JsonResponse({
            "ok":               True,
            "message":          msg,
            "new_due_date":     new_due_date.isoformat(),
            "renewal_count":    txn.renewal_count,
            "max_renewals":     max_renewals,
            "late_fine_amount": str(late_fine_created.amount) if late_fine_created else "0.00",
        })

    messages.success(request, msg)

    # ── Email: renewal confirmation ────────────────────────────────────────────
    if _member_emails_on(library) and getattr(member, "email", None):
        _send_email(
            "send_book_renewed_email",
            member,
            txn,
            late_fine_created.amount if late_fine_created else None,
        )

    return redirect("transactions:transaction_detail", pk=pk)


# ─────────────────────────────────────────────────────────────────────────────
# 6. Overdue List
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def overdue_list(request):
    library = _get_library_or_404(request)
    guard = _require_active_subscription(request, library)
    if guard: return guard
    _sync_overdue_if_stale(library)

    qs = (
        Transaction.objects.for_library(library)
        .filter(status=Transaction.STATUS_OVERDUE)
        .select_related("member", "book")
    )

    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(
            Q(member__first_name__icontains=q)
            | Q(member__last_name__icontains=q)
            | Q(book__title__icontains=q)
        )

    severity = request.GET.get("severity", "")
    today    = date.today()
    if severity == "mild":
        qs = qs.filter(due_date__gte=today - timedelta(days=7))
    elif severity == "moderate":
        qs = qs.filter(
            due_date__range=(today - timedelta(days=30), today - timedelta(days=8))
        )
    elif severity == "severe":
        qs = qs.filter(due_date__lt=today - timedelta(days=30))

    overdue_transactions = list(qs.order_by("due_date"))

    sort = request.GET.get("sort", "")
    if sort == "overdue_days":
        overdue_transactions.sort(key=lambda t: t.overdue_days)
    elif sort == "-fine_amount":
        overdue_transactions.sort(key=lambda t: t.fine_amount, reverse=True)
    else:
        overdue_transactions.sort(key=lambda t: t.overdue_days, reverse=True)

    total_fine = sum(t.fine_amount for t in overdue_transactions)

    return render(request, "transactions/overdue_list.html", {
        "overdue_transactions": overdue_transactions,
        "total_fine":           total_fine,
    })


# ─────────────────────────────────────────────────────────────────────────────
# 7. Fine List
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def fine_list(request):
    library = _get_library_or_404(request)
    guard = _require_active_subscription(request, library)
    if guard: return guard
    qs = (
        Fine.objects.for_library(library)
        .select_related("transaction__member", "transaction__book")
    )

    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(
            Q(transaction__member__first_name__icontains=q)
            | Q(transaction__member__last_name__icontains=q)
            | Q(transaction__pk__icontains=q)
        )

    status = request.GET.get("status", "").strip()
    if status in (Fine.STATUS_UNPAID, Fine.STATUS_PAID, Fine.STATUS_WAIVED):
        qs = qs.filter(status=status)

    fine_type = request.GET.get("type", "").strip()
    if fine_type in (Fine.TYPE_OVERDUE, Fine.TYPE_LOST, Fine.TYPE_DAMAGE):
        qs = qs.filter(fine_type=fine_type)

    date_from = request.GET.get("date_from", "")
    date_to   = request.GET.get("date_to",   "")
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    base = Fine.objects.for_library(library)

    def _agg(subqs):
        return subqs.aggregate(t=Sum("amount"))["t"] or Decimal("0.00")

    total_fine       = _agg(base)
    total_fine_count = base.count()
    unpaid_qs        = base.filter(status=Fine.STATUS_UNPAID)
    unpaid_fine      = _agg(unpaid_qs)
    unpaid_count     = unpaid_qs.count()
    paid_qs          = base.filter(status=Fine.STATUS_PAID)
    paid_fine        = _agg(paid_qs)
    paid_count       = paid_qs.count()

    paginator = Paginator(qs.order_by("-created_at"), 25)
    fines     = paginator.get_page(request.GET.get("page", 1))

    return render(request, "transactions/fine_list.html", {
        "fines":            fines,
        "total_fine":       total_fine,
        "total_fine_count": total_fine_count,
        "unpaid_fine":      unpaid_fine,
        "unpaid_count":     unpaid_count,
        "paid_fine":        paid_fine,
        "paid_count":       paid_count,
    })


# ─────────────────────────────────────────────────────────────────────────────
# 8. Mark Fine Paid  (POST only)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def mark_fine_paid(request, pk=None):
    if request.method != "POST":
        return redirect("transactions:fine_list")

    library = _get_library_or_404(request)
    guard = _require_active_subscription(request, library)
    if guard: return guard
    form    = MarkFinePaidForm(request.POST)

    if form.is_valid():
        cd   = form.cleaned_data
        fine = get_object_or_404(
            Fine.objects.for_library(library).select_related("transaction"),
            fine_id=cd["fine_id"],
        )
        # ── Rule: partial payments require the toggle to be ON ──────────────
        payment_amount = cd.get("payment_amount", None)
        is_partial = (
            payment_amount is not None
            and payment_amount < fine.amount
        )
        if is_partial and not _partial_payment_allowed(library):
            messages.error(
                request,
                f"Partial payment of ₹{payment_amount} is not allowed. "
                "Enable partial payments in Settings → Fine & Loans → Allow Partial Payment, "
                f"or pay the full amount of ₹{fine.amount}.",
            )
            return redirect(request.POST.get("next") or reverse("transactions:fine_list"))

        if fine.status == Fine.STATUS_UNPAID:
            with db_transaction.atomic():
                fine.mark_paid(
                    method=cd["payment_method"],
                    ref=cd.get("payment_ref", ""),
                )
                # Re-fetch the transaction from DB so _reset sees fresh state
                # (fine.transaction is a cached object; mark_paid() only updated
                #  the Fine row — the Transaction object in memory is stale)
                txn_fresh = Transaction.objects.select_related("member").get(
                    pk=fine.transaction_id
                )
                # Flip overdue → overdue_settled if all fines now paid/waived
                _reset_overdue_after_fine_settled(txn_fresh, library)
            messages.success(request, f"Fine of ₹{fine.amount} marked as paid.")
            # ── Email: fine payment receipt ────────────────────────────────────
            _member = fine.transaction.member if hasattr(fine.transaction, "member") else None
            if _member and _member_emails_on(library) and getattr(_member, "email", None):
                _send_email("send_fine_paid_email", _member, fine, fine.transaction)
        else:
            messages.warning(request, "Fine is already paid.")
    else:
        messages.error(request, "Invalid form submission.")

    return redirect(request.POST.get("next") or reverse("transactions:fine_list"))


# ─────────────────────────────────────────────────────────────────────────────
# 9. Missing / Lost Books  (was 10)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def missing_books(request):
    library = _get_library_or_404(request)
    guard = _require_active_subscription(request, library)
    if guard: return guard
    qs = (
        MissingBook.objects.for_library(library)
        .select_related("transaction__member", "book")
    )

    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(
            Q(book__title__icontains=q)
            | Q(book__isbn__icontains=q)
            | Q(transaction__member__first_name__icontains=q)
            | Q(transaction__member__last_name__icontains=q)
        )

    status = request.GET.get("status", "").strip()
    if status in (MissingBook.STATUS_MISSING, MissingBook.STATUS_LOST, MissingBook.STATUS_RECOVERED):
        qs = qs.filter(status=status)

    penalty = request.GET.get("penalty", "").strip()
    if penalty == "pending":
        qs = qs.filter(penalty_amount__gt=0, penalty_paid=False)
    elif penalty == "paid":
        qs = qs.filter(penalty_paid=True)

    base            = MissingBook.objects.for_library(library)
    lost_count      = base.filter(status=MissingBook.STATUS_LOST).count()
    missing_count   = base.filter(status=MissingBook.STATUS_MISSING).count()
    recovered_count = base.filter(status=MissingBook.STATUS_RECOVERED).count()
    total_penalty   = base.aggregate(t=Sum("penalty_amount"))["t"] or Decimal("0.00")

    return render(request, "transactions/missing_books.html", {
        "missing_books":   qs.order_by("-created_at"),
        "lost_count":      lost_count,
        "missing_count":   missing_count,
        "recovered_count": recovered_count,
        "total_penalty":   total_penalty,
    })


# ─────────────────────────────────────────────────────────────────────────────
# 11. Mark Lost  (POST only)
# ─────────────────────────────────────────────────────────────────────────────
#
# RULE: Book lost → Fine(TYPE_LOST) created automatically
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def mark_lost(request, pk=None):
    if request.method != "POST":
        return redirect("transactions:missing_books")

    library = _get_library_or_404(request)
    guard = _require_active_subscription(request, library)
    if guard: return guard
    form    = MarkLostForm(request.POST)

    if not form.is_valid():
        messages.error(request, "Invalid submission.")
        return redirect("transactions:missing_books")

    txn = get_object_or_404(
        Transaction.objects.for_library(library).select_related("book"),
        pk=form.cleaned_data["transaction_id"],
    )

    if txn.status == Transaction.STATUS_RETURNED:
        messages.error(request, "Cannot mark a returned transaction as lost.")
        return redirect("transactions:missing_books")

    with db_transaction.atomic():
        fine_paid_now  = form.cleaned_data.get("fine_paid_now", False)
        penalty_amount = form.cleaned_data.get("book_price") or txn.fine_amount or Decimal("0.00")

        txn.status    = Transaction.STATUS_LOST
        txn.lost_date = date.today()
        txn.notes     = form.cleaned_data.get("notes", "")
        if fine_paid_now:
            txn.fine_paid      = True
            txn.fine_paid_date = date.today()
        txn.save(update_fields=["status", "lost_date", "notes", "fine_paid", "fine_paid_date", "updated_at"])

        book = txn.book
        # book.total_copies     = max(0, book.total_copies - 1)
        # book.available_copies = min(book.available_copies, book.total_copies)
        book.save(update_fields=["total_copies", "available_copies"])

        # Mark the physical copy as lost (blocks re-issue)
        from books.models import BookCopy as _BookCopy
        try:
            copy_to_lose = txn.book_copy or (
                _BookCopy.objects
                .filter(book=book, status__in=(_BookCopy.Status.BORROWED, "borrowed", "issued"))
                .order_by("updated_at")
                .first()
            )
            if copy_to_lose:
                copy_to_lose.status = _BookCopy.Status.LOST
                copy_to_lose.save(update_fields=["status", "updated_at"])
        except Exception:
            pass

        missing_obj, _ = MissingBook.objects.update_or_create(
            transaction=txn,
            defaults={
                "library":        library,
                "book":           book,
                "status":         MissingBook.STATUS_LOST,
                "reported_date":  date.today(),
                "notes":          form.cleaned_data.get("notes", ""),
                "penalty_amount": penalty_amount,
                "penalty_reason": form.cleaned_data.get("reason", "lost"),
                "penalty_paid":   fine_paid_now,
            },
        )

        # ── Rule 4/7: fine for lost book (only if auto_fine enabled) ──────────
        if _auto_fine_enabled(library):
            fine_obj, created = Fine.objects.get_or_create(
                library     = library,
                transaction = txn,
                fine_type   = Fine.TYPE_LOST,
                defaults    = {
                    "amount":    penalty_amount,
                    "status":    Fine.STATUS_PAID if fine_paid_now else Fine.STATUS_UNPAID,
                    "paid_date": date.today() if fine_paid_now else None,
                },
            )
            # If row already existed and we're paying now — update it
            if not created and fine_paid_now and fine_obj.status == Fine.STATUS_UNPAID:
                fine_obj.amount    = penalty_amount
                fine_obj.status    = Fine.STATUS_PAID
                fine_obj.paid_date = date.today()
                fine_obj.save(update_fields=["amount", "status", "paid_date", "updated_at"])

            # Also sweep any other unpaid fines on this txn if paying now
            if fine_paid_now:
                Fine.objects.for_library(library).filter(
                    transaction=txn,
                    status=Fine.STATUS_UNPAID,
                ).exclude(fine_type=Fine.TYPE_LOST).update(
                    status=Fine.STATUS_PAID, paid_date=date.today()
                )

    paid_msg = f" Penalty of ₹{penalty_amount} collected." if fine_paid_now else f" Penalty of ₹{penalty_amount} recorded (unpaid)."
    messages.success(request, f'"{txn.book.title}" marked as lost. Copy removed from inventory.' + paid_msg)

    # ── Email: book marked lost notification ───────────────────────────────────
    _member = txn.member if hasattr(txn, "member") else None
    if _member and _member_emails_on(library) and getattr(_member, "email", None):
        _send_email("send_book_lost_email", _member, txn)

    return redirect("transactions:missing_books")


# ─────────────────────────────────────────────────────────────────────────────
# 12. Mark Recovered  (POST only)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def mark_recovered(request, pk):
    if request.method != "POST":
        return redirect("transactions:missing_books")

    library = _get_library_or_404(request)
    guard = _require_active_subscription(request, library)
    if guard: return guard
    missing = get_object_or_404(
        MissingBook.objects.for_library(library).select_related("book"),
        pk=pk,
    )

    if missing.status == MissingBook.STATUS_RECOVERED:
        messages.info(request, "Book is already marked as recovered.")
        return redirect("transactions:missing_books")

    with db_transaction.atomic():
        missing.status = MissingBook.STATUS_RECOVERED
        missing.save(update_fields=["status", "updated_at"])

        book = missing.book
        book.total_copies     += 1
        book.available_copies += 1
        book.save(update_fields=["total_copies", "available_copies"])

    messages.success(request, f'"{missing.book.title}" marked as recovered.')
    return redirect("transactions:missing_books")


# ─────────────────────────────────────────────────────────────────────────────
# 13. Add / Update Penalty  (POST only)
# ─────────────────────────────────────────────────────────────────────────────
#
# RULE: Book lost → Fine(TYPE_LOST) upserted with the confirmed penalty amount
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def add_penalty(request, pk=None):
    if request.method != "POST":
        return redirect("transactions:missing_books")

    library = _get_library_or_404(request)
    guard = _require_active_subscription(request, library)
    if guard: return guard
    form    = AddPenaltyForm(request.POST)

    if not form.is_valid():
        messages.error(request, "Invalid penalty form.")
        return redirect("transactions:missing_books")

    cd      = form.cleaned_data
    missing = get_object_or_404(MissingBook.objects.for_library(library), pk=cd["missing_id"])

    with db_transaction.atomic():
        missing.penalty_amount = cd["penalty_amount"]
        missing.penalty_reason = cd["penalty_reason"]
        if cd.get("notes"):
            missing.notes = cd["notes"]
        missing.save(update_fields=["penalty_amount", "penalty_reason", "notes", "updated_at"])

        # ── Rule: upsert Fine(TYPE_LOST) with confirmed penalty amount ─────────
        Fine.objects.update_or_create(
            library     = library,
            transaction = missing.transaction,
            fine_type   = Fine.TYPE_LOST,
            defaults    = {"amount": cd["penalty_amount"], "status": Fine.STATUS_UNPAID},
        )

    messages.success(request, f"Penalty of ₹{cd['penalty_amount']} applied.")

    # ── Email: fine applied notification ──────────────────────────────────────
    if _member_emails_on(library):
        try:
            _txn    = missing.transaction
            _member = _txn.member if hasattr(_txn, "member") else None
            _fine   = Fine.objects.for_library(library).filter(
                transaction=_txn, fine_type=Fine.TYPE_LOST
            ).order_by("-created_at").first()
            if _member and _fine and getattr(_member, "email", None):
                _send_email("send_fine_created_email", _member, _fine, _txn)
        except Exception:
            pass

    return redirect("transactions:missing_books")


# ─────────────────────────────────────────────────────────────────────────────
# 14. Member Search API  (AJAX / JSON)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def member_search_api(request):
    library = _get_library_or_404(request)
    guard = _require_active_subscription(request, library)
    if guard: return guard
    owner   = library.user

    from members.models import Member

    q  = request.GET.get("q", "").strip()
    qs = Member.objects.filter(owner=owner, status="active")
    if q:
        qs = qs.filter(
            Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(member_id__icontains=q)
            | Q(email__icontains=q)
        )

    members_list = list(qs.order_by("first_name", "last_name")[:20])

    active_counts = {
        row["member_id"]: row["cnt"]
        for row in Transaction.objects.for_library(library).filter(
            member__in=members_list,
            status__in=(Transaction.STATUS_ISSUED, Transaction.STATUS_OVERDUE),
        ).values("member_id").annotate(cnt=Count("id"))
    }

    def _photo_url(member):
        for _f in ("photo", "profile_photo", "avatar", "profile_picture", "image", "profile_image"):
            _v = getattr(member, _f, None)
            if _v:
                try:
                    if hasattr(_v, "name") and _v.name:
                        return request.build_absolute_uri(_v.url)
                except Exception:
                    pass
        try:
            return request.build_absolute_uri(reverse("transactions:member_photo_image", args=[member.pk]))
        except Exception:
            return None

    results = [
        {
            "id":           m.pk,
            "name":         f"{m.first_name} {m.last_name}".strip(),
            "member_id":    m.member_id,
            "active_loans": active_counts.get(m.pk, 0),
            "borrow_limit": _get_borrow_limit(library, m),
            "photo_url":    _photo_url(m),
        }
        for m in members_list
    ]
    return JsonResponse({"results": results})


# ─────────────────────────────────────────────────────────────────────────────
# 15. Book Search API  (AJAX / JSON)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def book_search_api(request):
    try:
        library = _get_library_or_404(request)
        guard = _require_active_subscription(request, library)
        if guard: return guard
        owner   = library.user

        from books.models import Book, BookCopy as _BookCopy

        q = request.GET.get("q", "").strip()

        if q:
            matching_pks = (
                _BookCopy.objects
                .filter(book__owner=owner, status="available", copy_id__icontains=q)
                .values_list("book_id", flat=True).distinct()
            )
            qs = Book.objects.filter(owner=owner, available_copies__gt=0, pk__in=matching_pks)
        else:
            qs = Book.objects.filter(owner=owner, available_copies__gt=0)

        results = []
        for b in qs.order_by("title")[:20]:
            copy_filter = {"book": b, "status": "available"}
            if q:
                copy_filter["copy_id__icontains"] = q
            copy_ids = list(
                _BookCopy.objects.filter(**copy_filter)
                .values_list("copy_id", flat=True).order_by("copy_id")[:10]
            )
            results.append({
                "id":               b.pk,
                "title":            b.title,
                "author":           b.author,
                "isbn":             getattr(b, "isbn", "") or "",
                "book_id":          getattr(b, "book_id", "") or "",
                "available_copies": b.available_copies,
                "copy_ids":         copy_ids,
            })
        return JsonResponse({"results": results})

    except Exception as exc:
        import logging, traceback
        logging.getLogger("transactions.views").error("book_search_api: %s", traceback.format_exc())
        return JsonResponse({"results": [], "error": str(exc)}, status=200)


# ─────────────────────────────────────────────────────────────────────────────
# 16. Member Lookup  (AJAX / JSON)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def member_suggestions_api(request):
    """
    Autocomplete endpoint — returns up to 10 members whose member_id,
    first_name, or last_name contains the query string (case-insensitive).
    Called on every keystroke in the Member ID field.
    """
    library = _get_library_or_404(request)
    guard = _require_active_subscription(request, library)
    if guard: return guard
    owner   = library.user

    from members.models import Member

    q = request.GET.get("q", "").strip()
    if not q:
        return JsonResponse({"results": []})

    # Evaluate queryset into a list so we can reuse it without a second DB hit
    members = list(
        Member.objects.filter(owner=owner).filter(
            Q(member_id__icontains=q)
            | Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
        ).order_by("member_id")[:10]
    )

    if not members:
        return JsonResponse({"results": []})

    # Bulk unpaid fine totals — keyed by member PK
    member_pks = [m.pk for m in members]
    fine_totals = {
        row["transaction__member_id"]: row["total"]
        for row in Fine.objects.for_library(library)
        .filter(transaction__member_id__in=member_pks, status=Fine.STATUS_UNPAID)
        .values("transaction__member_id")
        .annotate(total=Sum("amount", output_field=_DF()))
    }

    def _suggestion_photo_url(member):
        if getattr(member, "photo", None):
            try:
                return request.build_absolute_uri(
                    reverse("transactions:member_photo_image", args=[member.pk])
                )
            except Exception:
                pass
        return None

    results = [
        {
            "member_id": m.member_id,
            "name":      f"{m.first_name} {m.last_name}".strip(),
            "status":    m.status,
            "total_due": str(fine_totals.get(m.pk, Decimal("0.00"))),
            "photo_url": _suggestion_photo_url(m),
        }
        for m in members
    ]

    return JsonResponse({"results": results})


@login_required
def member_lookup_api(request):
    library = _get_library_or_404(request)
    guard = _require_active_subscription(request, library)
    if guard: return guard
    owner   = library.user

    from members.models import Member

    raw_id = request.GET.get("member_id", "").strip().upper()
    if not raw_id:
        return JsonResponse({"found": False, "error": "No member_id supplied."})

    try:
        member = Member.objects.get(member_id=raw_id, owner=owner)
    except Member.DoesNotExist:
        return JsonResponse({"found": False, "error": f'Member ID "{raw_id}" not found.'})

    borrow_limit = _get_borrow_limit(library, member)
    active_loans = (
        Transaction.objects.for_library(library)
        .filter(member=member, status__in=(Transaction.STATUS_ISSUED, Transaction.STATUS_OVERDUE))
        .count()
    )
    total_due = (
        Fine.objects.for_library(library)
        .filter(transaction__member=member, status=Fine.STATUS_UNPAID)
        .aggregate(t=Sum("amount", output_field=_DF()))["t"] or Decimal("0.00")
    )
    slots = max(0, borrow_limit - active_loans) if borrow_limit else -1

    # photo is a raw BLOB — use member_photo_image route only when data exists
    photo_url = None
    if getattr(member, "photo", None):
        try:
            photo_url = request.build_absolute_uri(
                reverse("transactions:member_photo_image", args=[member.pk])
            )
        except Exception:
            pass

    return JsonResponse({
        "found":        True,
        "pk":           member.pk,
        "member_id":    member.member_id,
        "name":         f"{member.first_name} {member.last_name}".strip(),
        "role":         member.get_role_display() if hasattr(member, "get_role_display") else "",
        "email":        member.email or "",
        "status":       member.status,
        "active_loans": active_loans,
        "borrow_limit": borrow_limit,
        "slots":        slots,
        "total_due":    str(total_due),
        "photo_url":    photo_url,
    })


# ─────────────────────────────────────────────────────────────────────────────
# 17. Book Lookup  (AJAX / JSON)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def book_lookup_api(request):
    import traceback
    try:
        library = _get_library_or_404(request)
        guard = _require_active_subscription(request, library)
        if guard: return guard
        owner   = library.user

        from books.models import Book, BookCopy

        raw_id = request.GET.get("book_id", "").strip()
        if not raw_id:
            return JsonResponse({"found": False, "error": "No book_id supplied."})

        try:
            copy = BookCopy.objects.select_related("book").get(copy_id=raw_id, book__owner=owner)
        except BookCopy.DoesNotExist:
            return JsonResponse({"found": False, "error": f'Book copy ID "{raw_id}" not found.'})
        except Exception as e:
            return JsonResponse({"found": False, "error": f"[copy lookup] {type(e).__name__}: {e}"})

        book = copy.book

        try:
            available_copies = BookCopy.objects.filter(book=book, status="available").count()
            total_copies     = BookCopy.objects.filter(book=book).count()
        except Exception:
            available_copies = getattr(book, "available_copies", 0) or 0
            total_copies     = getattr(book, "total_copies", available_copies) or available_copies

        try:
            category_name = book.category.name if getattr(book, "category_id", None) else ""
        except Exception:
            category_name = ""

        cover_url = None
        try:
            if getattr(book, "cover_image", None):
                cover_url = request.build_absolute_uri(
                    reverse("transactions:book_cover_image", args=[book.pk])
                )
        except Exception:
            pass

        return JsonResponse({
            "found":            True,
            "pk":               book.pk,
            "copy_pk":          copy.pk,
            "book_id":          raw_id,
            "title":            getattr(book, "title",  "") or "",
            "author":           getattr(book, "author", "") or "",
            "isbn":             getattr(book, "isbn",   "") or "",
            "available_copies": available_copies,
            "total_copies":     total_copies,
            "category":         category_name,
            "cover_url":        cover_url,
            "copy_status":      getattr(copy, "status",    ""),
            "copy_condition":   getattr(copy, "condition", ""),
        })

    except Exception as exc:
        import logging
        logging.getLogger("transactions.views").error(
            "book_lookup_api: %s", traceback.format_exc()
        )
        return JsonResponse(
            {"found": False, "error": f"Server error: {type(exc).__name__}: {exc}"},
            status=200,
        )


# ─────────────────────────────────────────────────────────────────────────────
# 18. Book Cover Image  (serves BLOB by book PK)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def book_cover_image(request, pk):
    """
    Serve the cover image BLOB for a book identified by its PK.

    Owner is resolved via library.user (same pattern as every other view)
    so staff / superuser accounts that are the library admin still match.
    Returns 404 when the book has no cover — the <img> onerror handler
    in the template then shows the letter-initial fallback.
    """
    from books.models import Book

    library = _get_library_or_404(request)
    guard = _require_active_subscription(request, library)
    if guard: return guard
    owner   = library.user

    book = get_object_or_404(Book, pk=pk, owner=owner)

    raw = book.cover_image
    if not raw:
        raise Http404("No cover image.")

    image_bytes = bytes(raw)
    if not image_bytes:
        raise Http404("Cover image is empty.")

    mime     = (getattr(book, "cover_mime_type", None) or "image/jpeg").strip() or "image/jpeg"
    response = HttpResponse(image_bytes, content_type=mime)
    response["Cache-Control"] = "private, max-age=3600"
    return response


# ─────────────────────────────────────────────────────────────────────────────
# 18a. Book Cover by Copy ID
#      GET /transactions/api/book-cover/copy/<copy_id>/
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def book_cover_by_copy_id(request, copy_id):
    """
    Resolve a BookCopy.copy_id → Book → cover BLOB and serve it.
    """
    from books.models import BookCopy

    library = _get_library_or_404(request)
    guard = _require_active_subscription(request, library)
    if guard: return guard
    owner   = library.user

    try:
        copy = (
            BookCopy.objects
            .select_related("book")
            .get(copy_id=copy_id.strip().upper(), book__owner=owner)
        )
    except BookCopy.DoesNotExist:
        raise Http404(f"Copy ID '{copy_id}' not found.")

    book        = copy.book
    raw         = book.cover_image
    if not raw:
        raise Http404("This book has no cover image.")
    image_bytes = bytes(raw)
    if not image_bytes:
        raise Http404("Cover image is empty.")
    mime        = (getattr(book, "cover_mime_type", None) or "image/jpeg").strip() or "image/jpeg"
    response    = HttpResponse(image_bytes, content_type=mime)
    response["Cache-Control"] = "private, max-age=3600"
    return response

# ─────────────────────────────────────────────────────────────────────────────
# 19. Member Photo  (serves BLOB by member PK)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def member_photo_image(request, pk):
    """
    Serve the photo BLOB for a member identified by PK.
    photo_mime_type column is used when present, falls back to image/jpeg.
    Returns 404 when the member has no photo so the frontend initial-avatar
    fallback renders instead of a broken <img>.
    """
    from members.models import Member

    library = _get_library_or_404(request)
    guard = _require_active_subscription(request, library)
    if guard: return guard
    owner   = library.user

    member = get_object_or_404(Member, pk=pk, owner=owner)

    raw = getattr(member, "photo", None)
    if not raw:
        raise Http404("No photo.")

    image_bytes = bytes(raw)
    if not image_bytes:
        raise Http404("Photo is empty.")

    mime = (getattr(member, "photo_mime_type", None) or "image/jpeg").strip() or "image/jpeg"
    response = HttpResponse(image_bytes, content_type=mime)
    response["Cache-Control"] = "private, max-age=3600"
    return response
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import Transaction


@login_required
def export_transactions_excel(request):
    library = _get_library_or_404(request)
    guard = _require_plan_feature(request, library, "export", "Excel Export")
    if guard: return guard

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Transactions"

    headers = [
        "Transaction ID",
        "Member",
        "Book",
        "Issue Date",
        "Due Date",
        "Return Date",
        "Status",
        "Days Borrowed",
        "Overdue Days",
        "Fine Amount",
        "Damage Charge",
        "Fine Paid",
        "Issued By",
        "Returned To",
    ]

    # Header Style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    sheet.append(headers)

    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    transactions = Transaction.objects.for_library(library).select_related("member", "book")

    row = 2
    for t in transactions:

        # Fine paid logic
        if t.status == "issued" or t.fine_amount == 0:
            fine_status = ""
        else:
            fine_status = "✔" if t.fine_paid else "❌"

        sheet.append([
            t.transaction_id,
            str(t.member),
            str(t.book),
            t.issue_date,
            t.due_date,
            t.return_date,
            t.status,
            t.days_borrowed,
            t.overdue_days,
            float(t.fine_amount),
            float(t.damage_charge),
            fine_status,
            t.issued_by,
            t.returned_to,
        ])

        # STATUS COLOR
        status_cell = sheet.cell(row=row, column=7)

        if t.status == "returned":
            status_cell.font = Font(color="00B050", bold=True)  # green
        elif t.status == "lost":
            status_cell.font = Font(color="FF0000", bold=True)  # red
        elif t.status == "issued":
            status_cell.font = Font(color="0070C0", bold=True)  # blue

        # FINE PAID COLOR
        fine_cell = sheet.cell(row=row, column=12)

        if fine_status == "✔":
            fine_cell.font = Font(color="00B050", bold=True)
        elif fine_status == "❌":
            fine_cell.font = Font(color="FF0000", bold=True)

        row += 1

    # Column widths
    sheet.column_dimensions["A"].width = 27
    sheet.column_dimensions["B"].width = 27
    sheet.column_dimensions["C"].width = 70

    for col in range(4, 15):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 18

    # Freeze header
    sheet.freeze_panes = "A2"

    # Filter
    sheet.auto_filter.ref = sheet.dimensions

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = "attachment; filename=Transactions.xlsx"

    workbook.save(response)

    return response