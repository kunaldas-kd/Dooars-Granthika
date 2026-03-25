"""
books/views.py
Dooars Granthika — Books module views.
"""

import io
from datetime import date

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from .forms import BookForm
from .models import Book, BookCopy, Category
from .services import create_book_copies

# ── How many available copies counts as "low stock"
LOW_STOCK_THRESHOLD = 3


# ─────────────────────────────────────────────────────────────
# Internal helpers
# ─────────────────────────────────────────────────────────────

def _user_books(user):
    return Book.objects.select_related("category").filter(owner=user)


def _user_categories(user):
    return Category.objects.filter(owner=user)


def _filter_books(qs, request):
    q        = request.GET.get("q", "").strip()
    category = request.GET.get("category", "").strip()
    stock    = request.GET.get("stock", "").strip()

    if q:
        qs = qs.filter(
            Q(title__icontains=q)    |
            Q(author__icontains=q)   |
            Q(isbn__icontains=q)     |
            Q(publisher__icontains=q)
        )
    if category:
        qs = qs.filter(category__slug=category)

    if stock in ("available", "low-stock", "out-stock"):
        qs = qs.annotate(
            _avail=Count(
                "copies",
                filter=Q(copies__status=BookCopy.Status.AVAILABLE),
            )
        )
        if stock == "available":
            qs = qs.filter(_avail__gt=LOW_STOCK_THRESHOLD)
        elif stock == "low-stock":
            qs = qs.filter(_avail__gt=0, _avail__lte=LOW_STOCK_THRESHOLD)
        elif stock == "out-stock":
            qs = qs.filter(_avail=0)

    return qs


def _get_library_code(user) -> str:
    """
    Return the 3-character library code for the given user.
    Reads accounts_library.library_name — first 3 characters, uppercased.

    e.g.  library_name = "Dooars Granthika"  →  "DOO"
    """
    from .services import derive_library_code
    try:
        library = getattr(user, "library", None)
        if library is None:
            from accounts.models import Library
            library = Library.objects.get(user=user)
        return derive_library_code(library)
    except Exception as exc:
        raise ValueError(
            f"Could not derive library code from library_name: {exc}. "
            "Ensure accounts_library.library_name is set in Settings."
        )


def _find_title_author_edition_duplicate(user, title, author, edition):
    return Book.objects.filter(
        owner=user,
        title__iexact=title,
        author__iexact=author,
        edition__iexact=edition,
    ).first()


# ─────────────────────────────────────────────────────────────
# Book List
# ─────────────────────────────────────────────────────────────

@login_required
def book_list(request):
    qs = _user_books(request.user)
    qs = _filter_books(qs, request)

    paginator = Paginator(qs, 20)
    page_obj  = paginator.get_page(request.GET.get("page"))

    return render(request, "books/book_list.html", {
        "books":      page_obj,
        "page_obj":   page_obj,
        "paginator":  paginator,
        "categories": _user_categories(request.user),
    })


# ─────────────────────────────────────────────────────────────
# Book Detail
# ─────────────────────────────────────────────────────────────

@login_required
def book_detail(request, pk):
    book = get_object_or_404(
        _user_books(request.user).prefetch_related("copies"),
        pk=pk,
    )
    return render(request, "books/book_detail.html", {"book": book})


# ─────────────────────────────────────────────────────────────
# Book Create
# ─────────────────────────────────────────────────────────────

@login_required
def book_create(request):
    from .forms import ExcelImportForm, parse_excel_rows

    form_type = request.POST.get("form_type", "manual") if request.method == "POST" else "manual"

    if request.method == "POST" and form_type == "import_upload":
        import_form = ExcelImportForm(request.POST, request.FILES)
        if import_form.is_valid():
            results = parse_excel_rows(import_form.cleaned_data["excel_file"], request.user)
            if not results:
                messages.error(request, "The file appears to be empty.")
            else:
                session_rows = []
                for r in results:
                    d = dict(r["data"])
                    cat = d.get("category")
                    d["category_pk"]       = cat.pk   if cat else None
                    d["category_name"]     = cat.name if cat else ""
                    d["_category_created"] = d.get("_category_created", False)
                    d.pop("category", None)
                    # Decimal is not JSON-serialisable — store as string
                    if d.get("price") is not None:
                        d["price"] = str(d["price"])
                    session_rows.append({
                        "row":    r["row"],
                        "data":   d,
                        "status": r["status"],
                        "errors": r["errors"],
                        "book_pk": r["book"].pk if r["book"] else None,
                    })
                request.session["import_preview"] = session_rows
                new_c = sum(1 for r in results if r["status"] == "new")
                dup_c = sum(1 for r in results if r["status"] == "duplicate")
                err_c = sum(1 for r in results if r["status"] == "error")
                return render(request, "books/book_form.html", {
                    "form":             BookForm(user=request.user),
                    "categories":       _user_categories(request.user),
                    "import_form":      import_form,
                    "import_step":      "preview",
                    "import_rows":      session_rows,
                    "import_new_count": new_c,
                    "import_dup_count": dup_c,
                    "import_err_count": err_c,
                })
        return render(request, "books/book_form.html", {
            "form":        BookForm(user=request.user),
            "categories":  _user_categories(request.user),
            "import_form": import_form,
            "import_step": None,
        })

    if request.method == "POST" and form_type == "import_confirm":
        session_rows  = request.session.pop("import_preview", [])
        selected_rows = set(request.POST.getlist("selected_rows"))
        created = updated = skipped = 0
        lib_code = _get_library_code(request.user)

        for r in session_rows:
            if str(r["row"]) not in selected_rows or r["status"] == "error":
                skipped += 1
                continue
            d   = r["data"]
            cat = Category.objects.filter(pk=d.get("category_pk")).first() if d.get("category_pk") else None
            total = int(d.get("total_copies") or 1)

            title   = d.get("title", "")
            author  = d.get("author", "")
            edition = d.get("edition", "")

            existing = _find_title_author_edition_duplicate(request.user, title, author, edition)
            if existing:
                with transaction.atomic():
                    create_book_copies(existing, lib_code, total)
                updated += 1
                continue

            with transaction.atomic():
                book = Book.objects.create(
                    owner            = request.user,
                    isbn             = d["isbn"],
                    title            = title,
                    author           = author,
                    category         = cat,
                    publisher        = d.get("publisher", ""),
                    publication_year = d.get("publication_year") or None,
                    language         = d.get("language", ""),
                    edition          = edition,
                    shelf_location   = d.get("shelf_location", ""),
                    description      = d.get("description", ""),
                    price            = d.get("price"),
                    total_copies     = total,
                    available_copies = total,
                )
                create_book_copies(book, lib_code, total)
            created += 1

        parts = []
        if created: parts.append(f"{created} book{'s' if created != 1 else ''} imported")
        if updated: parts.append(f"{updated} updated")
        if skipped: parts.append(f"{skipped} skipped")
        messages.success(request, " · ".join(parts) + ".")
        return redirect("books:book_list")

    if request.method == "POST" and form_type == "manual":
        form = BookForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            lib_code = _get_library_code(request.user)
            total    = form.cleaned_data.get("total_copies", 1)

            with transaction.atomic():
                book                  = form.save(commit=False)
                book.owner        = request.user
                book.category     = form.cleaned_data["category"]
                book.total_copies     = total
                book.available_copies = total
                book.price            = form.cleaned_data.get("price") or None
                img = form.cleaned_data.get("cover_image")
                if img:
                    book.cover_image     = img["data"]
                    book.cover_mime_type = img["mime"]
                book.save()
                create_book_copies(book, lib_code, total)

            if hasattr(form, "_created_category"):
                messages.info(request, f'New category "{form._created_category}" was created.')
            elif hasattr(form, "_reused_category"):
                messages.info(request, f'Existing category "{form._reused_category}" was reused.')
            messages.success(
                request,
                f'"{book.title}" added with {total} physical '
                f'{"copy" if total == 1 else "copies"}.',
            )
            return redirect("books:book_detail", pk=book.pk)
    else:
        form = BookForm(user=request.user)

    return render(request, "books/book_form.html", {
        "form":        form,
        "categories":  _user_categories(request.user),
        "import_form": __import__("books.forms", fromlist=["ExcelImportForm"]).ExcelImportForm(),
        "import_step": None,
    })


# ─────────────────────────────────────────────────────────────
# Book Update
# ─────────────────────────────────────────────────────────────

@login_required
def book_update(request, pk):
    book = get_object_or_404(Book, pk=pk, owner=request.user)

    if request.method == "POST":
        form = BookForm(request.POST, request.FILES, instance=book, user=request.user)
        if form.is_valid():
            lib_code  = _get_library_code(request.user)
            new_total = form.cleaned_data.get("total_copies", book.copy_count)

            with transaction.atomic():
                updated          = form.save(commit=False)
                updated.category = form.cleaned_data["category"]
                updated.price    = form.cleaned_data.get("price") or None
                img = form.cleaned_data.get("cover_image")
                cover_changed = False
                if img is False:
                    # User ticked "Clear" — wipe the stored image
                    updated.cover_image     = None
                    updated.cover_mime_type = ""
                    cover_changed = True
                elif img:
                    # New file uploaded — replace the stored image
                    updated.cover_image     = img["data"]
                    updated.cover_mime_type = img["mime"]
                    cover_changed = True
                else:
                    # No change — preserve the existing image
                    updated.cover_image     = book.cover_image
                    updated.cover_mime_type = book.cover_mime_type
                updated.save()

                # BinaryField is editable=False by default — Django's UPDATE
                # query omits it. Force-write cover columns whenever they changed.
                update_kwargs = {
                    "total_copies": book.copy_count,
                }
                if cover_changed:
                    update_kwargs["cover_image"]     = updated.cover_image
                    update_kwargs["cover_mime_type"] = updated.cover_mime_type
                Book.objects.filter(pk=book.pk).update(**update_kwargs)

                current_total = book.copy_count
                if new_total > current_total:
                    extra = new_total - current_total
                    create_book_copies(book, lib_code, extra)
                    messages.info(request, f"{extra} new physical {'copy' if extra == 1 else 'copies'} generated.")

            if hasattr(form, "_created_category"):
                messages.info(request, f'New category "{form._created_category}" was created.')
            elif hasattr(form, "_reused_category"):
                messages.info(request, f'Existing category "{form._reused_category}" was reused.')
            messages.success(request, f'"{book.title}" was updated successfully.')
            return redirect("books:book_detail", pk=book.pk)
    else:
        form = BookForm(instance=book, user=request.user,
                        initial={"total_copies": book.copy_count})

    return render(request, "books/book_form.html", {
        "form":       form,
        "categories": _user_categories(request.user),
    })


# ─────────────────────────────────────────────────────────────
# Book Delete
# ─────────────────────────────────────────────────────────────

@login_required
def book_delete(request, pk):
    book = get_object_or_404(Book, pk=pk, owner=request.user)
    if request.method == "POST":
        title = book.title
        book.delete()
        messages.success(request, f"{title} was deleted.")
        return redirect("books:book_list")
    return render(request, "books/book_delete.html", {"book": book})


# ─────────────────────────────────────────────────────────────
# Borrow / Return a Copy
# ─────────────────────────────────────────────────────────────

@login_required
def borrow_copy(request):
    if request.method != "POST":
        messages.error(request, "Invalid request method.")
        return redirect("books:book_list")

    copy_id = request.POST.get("copy_id", "").strip().upper()
    if not copy_id:
        messages.error(request, "No Copy ID provided.")
        return redirect("books:book_list")

    copy = get_object_or_404(
        BookCopy.objects.select_related("book"),
        copy_id=copy_id,
        book__owner=request.user,
    )

    try:
        copy.borrow()
        messages.success(request, f'Copy {copy_id} of "{copy.book.title}" has been marked as borrowed.')
    except ValueError as exc:
        messages.error(request, str(exc))

    return redirect("books:book_detail", pk=copy.book.pk)


@login_required
def return_copy(request):
    if request.method != "POST":
        messages.error(request, "Invalid request method.")
        return redirect("books:book_list")

    copy_id = request.POST.get("copy_id", "").strip().upper()
    if not copy_id:
        messages.error(request, "No Copy ID provided.")
        return redirect("books:book_list")

    copy = get_object_or_404(
        BookCopy.objects.select_related("book"),
        copy_id=copy_id,
        book__owner=request.user,
    )

    try:
        copy.return_copy()
        messages.success(request, f'Copy {copy_id} of "{copy.book.title}" has been returned — now available.')
    except ValueError as exc:
        messages.error(request, str(exc))

    return redirect("books:book_detail", pk=copy.book.pk)


# ─────────────────────────────────────────────────────────────
# Stock Dashboard
# ─────────────────────────────────────────────────────────────

@login_required
def stock_dashboard(request):
    all_books  = _user_books(request.user).prefetch_related("copies")
    total_books = all_books.count()

    low_stock_count    = 0
    out_of_stock_count = 0
    available_books    = 0
    for book in all_books:
        avail = book.available_copy_count
        if avail == 0:
            out_of_stock_count += 1
        elif avail <= LOW_STOCK_THRESHOLD:
            low_stock_count += 1
            available_books += 1
        else:
            available_books += 1

    def pct(n):
        return round(n / total_books * 100) if total_books else 0

    # Progress bar segments must be mutually exclusive and sum to 100.
    # above_threshold = books with > LOW_STOCK_THRESHOLD available copies
    above_threshold = total_books - low_stock_count - out_of_stock_count
    _low_pct  = pct(low_stock_count)
    _out_pct  = pct(out_of_stock_count)
    _above_pct = 100 - _low_pct - _out_pct  # absorbs rounding remainder

    category_stats = []
    for cat in _user_categories(request.user):
        cat_qs    = all_books.filter(category=cat)
        total     = cat_qs.count()
        available = sum(1 for b in cat_qs if b.available_copy_count > 0)
        category_stats.append({
            "name":      cat.name,
            "total":     total,
            "available": available,
            "pct":       round(available / total * 100) if total else 0,
        })

    most_issued_qs = sorted(all_books, key=lambda b: b.borrowed_copy_count, reverse=True)[:5]
    most_issued    = []
    for b in most_issued_qs:
        b.issue_count = b.borrowed_copy_count
        most_issued.append(b)

    recent_books   = all_books.order_by("-created_at")[:5]
    low_stock_list = [b for b in all_books if 0 < b.available_copy_count <= LOW_STOCK_THRESHOLD]

    return render(request, "books/book_stock_dashboard.html", {
        "total_books":        total_books,
        "available_books":    available_books,
        "low_stock_count":    low_stock_count,
        "out_of_stock_count": out_of_stock_count,
        # Stat-card percentages (available includes low-stock books too)
        "available_pct":      pct(available_books),
        "low_pct":            _low_pct,
        "out_pct":            _out_pct,
        # Progress-bar percentages — three exclusive segments summing to 100
        "bar_above_pct":      _above_pct,
        "bar_low_pct":        _low_pct,
        "bar_out_pct":        _out_pct,
        "category_stats":     category_stats,
        "most_issued":        most_issued,
        "recent_books":       recent_books,
        "low_stock_list":     low_stock_list,
        "low_threshold":      LOW_STOCK_THRESHOLD,
    })


# ─────────────────────────────────────────────────────────────
# Export
# ─────────────────────────────────────────────────────────────

@login_required
def export_books(request):
    qs = _user_books(request.user).prefetch_related("copies")
    qs = _filter_books(qs, request)

    all_books     = list(qs)
    total_count   = len(all_books)
    preview_limit = 8
    preview_books = all_books[:preview_limit]
    preview_more  = max(0, total_count - preview_limit)

    total_sum     = sum(b.copy_count           for b in all_books)
    available_sum = sum(b.available_copy_count for b in all_books)
    issued_sum    = total_sum - available_sum

    return render(request, "books/book_export.html", {
        "preview_books":        preview_books,
        "preview_more":         preview_more,
        "total_count":          total_count,
        "total_copies_sum":     total_sum,
        "available_copies_sum": available_sum,
        "issued_copies_sum":    issued_sum,
        "categories":           _user_categories(request.user),
    })


@login_required
def export_books_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, "openpyxl is not installed. Run: pip install openpyxl")
        return redirect("books:export_books")

    qs = _user_books(request.user).prefetch_related("copies")
    qs = _filter_books(qs, request)
    books = list(qs)  # evaluate once — prefetch cache used for all copy lookups

    HEADER_FILL  = PatternFill("solid", fgColor="0A1628")
    HEADER_FONT  = Font(color="FFFFFF", bold=True, size=9)
    TOTALS_FILL  = PatternFill("solid", fgColor="1E3A5F")
    TOTALS_FONT  = Font(color="FFFFFF", bold=True, size=9)
    GREEN_FILL   = PatternFill("solid", fgColor="DCFCE7")
    ORANGE_FILL  = PatternFill("solid", fgColor="FFF7ED")
    RED_FILL     = PatternFill("solid", fgColor="FEE2E2")
    CENTER       = Alignment(horizontal="center", vertical="center")
    THIN         = Side(style="thin", color="D1D5DB")
    BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Book Catalogue"

    headers    = ["#","Title","Author","ISBN","Category","Publisher","Language","Edition","Total Copies","Available","Borrowed","Price (₹)","Shelf Location","Added On"]
    col_widths = [4, 32, 22, 18, 16, 20, 11, 12, 12, 10, 10, 11, 14, 12]

    ws.append(headers)
    ws.row_dimensions[1].height = 22
    for ci, (_, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"

    for idx, book in enumerate(books, 1):
        avail    = book.available_copy_count
        borrowed = book.borrowed_copy_count
        total    = book.copy_count
        price    = float(book.price) if book.price is not None else ""
        ws.append([idx, book.title, book.author, book.isbn,
            book.category.name if book.category else "",
            book.publisher, book.language, book.edition,
            total, avail, borrowed, price,
            book.shelf_location, book.created_at.strftime("%d/%m/%Y")])
        r = ws.max_row
        avail_cell = ws.cell(row=r, column=10)
        if avail == 0:                          avail_cell.fill = RED_FILL
        elif avail <= LOW_STOCK_THRESHOLD:      avail_cell.fill = ORANGE_FILL
        else:                                   avail_cell.fill = GREEN_FILL
        for ci in range(1, len(headers) + 1):
            ws.cell(row=r, column=ci).border = BORDER

    last_data = ws.max_row
    if last_data >= 2:
        ws.append(["", "TOTALS", "", "", "", "", "", "",
            f"=SUM(I2:I{last_data})", f"=SUM(J2:J{last_data})",
            f"=SUM(K2:K{last_data})", "", "", ""])
    else:
        ws.append(["", "TOTALS", "", "", "", "", "", "", 0, 0, 0, "", "", ""])
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=ci)
        cell.fill = TOTALS_FILL; cell.font = TOTALS_FONT
        cell.alignment = CENTER; cell.border = BORDER
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_data}"

    ws2 = wb.create_sheet("Physical Copies")
    s2_headers = ["Copy ID","Book Title","Author","ISBN","Status","Borrowed At","Returned At","Created"]
    ws2.append(s2_headers)
    ws2.row_dimensions[1].height = 22
    col_widths2 = [18, 32, 22, 18, 12, 18, 18, 14]
    for ci, w in enumerate(col_widths2, 1):
        cell = ws2.cell(row=1, column=ci)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = BORDER
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A2"

    for book in books:  # reuse the same list — no second DB hit
        for copy in book.copies.all():
            ws2.append([copy.copy_id, book.title, book.author, book.isbn,
                copy.get_status_display(),
                copy.borrowed_at.strftime("%d/%m/%Y %H:%M") if copy.borrowed_at else "",
                copy.returned_at.strftime("%d/%m/%Y %H:%M") if copy.returned_at else "",
                copy.created_at.strftime("%d/%m/%Y")])
            for ci in range(1, len(s2_headers) + 1):
                ws2.cell(row=ws2.max_row, column=ci).border = BORDER

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    filename = f"dooars_granthika_books_{date.today():%Y%m%d}.xlsx"
    resp = HttpResponse(buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# ─────────────────────────────────────────────────────────────
# Update Stock / Cover / Import / Template
# ─────────────────────────────────────────────────────────────

@login_required
def update_stock(request):
    return render(request, "books/book_stock_update.html", {
        "categories": _user_categories(request.user),
    })


@login_required
def book_cover(request, pk):
    from django.http import Http404
    book = get_object_or_404(Book, pk=pk, owner=request.user)
    raw = book.cover_image
    if raw is None:
        raise Http404("No cover image.")
    image_bytes = bytes(raw)
    if not image_bytes:
        raise Http404("Cover image is empty.")

    # ETag from updated_at — changes whenever the book record is saved
    etag = f'"{pk}-{int(book.updated_at.timestamp())}"'
    if request.META.get("HTTP_IF_NONE_MATCH") == etag:
        return HttpResponse(status=304)

    mime = (book.cover_mime_type or "image/jpeg").strip() or "image/jpeg"
    response = HttpResponse(image_bytes, content_type=mime)
    response["ETag"] = etag
    response["Cache-Control"] = "private, max-age=0, must-revalidate"
    return response


@login_required
def import_books_excel(request):
    from .forms import ExcelImportForm, parse_excel_rows

    step = request.POST.get("step", "upload")

    if request.method == "POST" and step == "upload":
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            results = parse_excel_rows(form.cleaned_data["excel_file"], request.user)
            if not results:
                messages.error(request, "The file appears to be empty.")
                return render(request, "books/book_import.html", {"form": form, "step": "upload"})

            session_rows = []
            for r in results:
                d   = dict(r["data"])
                cat = d.get("category")
                d["category_pk"]   = cat.pk   if cat else None
                d["category_name"] = cat.name if cat else ""
                d.pop("category", None)
                d.pop("_category_created", None)
                # Decimal is not JSON-serialisable — store as string
                if d.get("price") is not None:
                    d["price"] = str(d["price"])
                session_rows.append({
                    "row": r["row"], "data": d, "status": r["status"],
                    "errors": r["errors"],
                    "book_pk":    r["book"].pk  if r["book"] else None,
                    "book_title": str(r["book"]) if r["book"] else "",
                })
            request.session["import_preview"] = session_rows
            return render(request, "books/book_import.html", {
                "step":      "preview",
                "rows":      session_rows,
                "new_count": sum(1 for r in results if r["status"] == "new"),
                "dup_count": sum(1 for r in results if r["status"] == "duplicate"),
                "err_count": sum(1 for r in results if r["status"] == "error"),
                "form":      ExcelImportForm(),
            })
        return render(request, "books/book_import.html", {"form": form, "step": "upload"})

    if request.method == "POST" and step == "confirm":
        session_rows  = request.session.pop("import_preview", [])
        if not session_rows:
            messages.error(request, "Session expired. Please re-upload the file.")
            return redirect("books:import_books_excel")

        selected_rows = set(request.POST.getlist("selected_rows"))
        lib_code      = _get_library_code(request.user)
        created_count = updated_count = skipped_count = 0

        for r in session_rows:
            if str(r["row"]) not in selected_rows or r["status"] == "error":
                skipped_count += 1
                continue
            d = r["data"]
            try:
                total = max(int(d.get("total_copies") or 1), 1)
            except (TypeError, ValueError):
                total = 1

            title   = d.get("title", "")
            author  = d.get("author", "")
            edition = d.get("edition", "")

            existing = _find_title_author_edition_duplicate(request.user, title, author, edition)
            if existing:
                with transaction.atomic():
                    create_book_copies(existing, lib_code, total)
                updated_count += 1
                continue

            cat = None
            if d.get("category_pk"):
                cat = Category.objects.filter(pk=d["category_pk"]).first()

            with transaction.atomic():
                if r["status"] == "duplicate":
                    book = Book.objects.filter(owner=request.user, isbn=d["isbn"]).first()
                    if book:
                        create_book_copies(book, lib_code, total)
                        updated_count += 1
                        continue

                book = Book.objects.create(
                    owner=request.user, isbn=d["isbn"],
                    title=title, author=author, category=cat,
                    publisher=d.get("publisher", ""),
                    publication_year=d.get("publication_year") or None,
                    language=d.get("language", ""), edition=edition,
                    shelf_location=d.get("shelf_location", ""),
                    total_copies=total,
                    available_copies=total,
                    description=d.get("description", ""),
                    price=d.get("price") or None,
                )
                create_book_copies(book, lib_code, total)
            created_count += 1

        parts = []
        if created_count: parts.append(f"{created_count} imported")
        if updated_count: parts.append(f"{updated_count} updated")
        if skipped_count: parts.append(f"{skipped_count} skipped")
        messages.success(request, " · ".join(parts) + ".")
        return redirect("books:book_list")

    return render(request, "books/book_import.html", {
        "form": __import__("books.forms", fromlist=["ExcelImportForm"]).ExcelImportForm(),
        "step": "upload",
    })


@login_required
def download_import_template(request):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        messages.error(request, "openpyxl is not installed.")
        return redirect("books:import_books_excel")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Books"

    # "Available Copies" removed — it always equals Total Copies on import.
    # "Price (₹)" added.
    headers    = ["Title","Author","ISBN","Category","Publisher","Publication Year","Language","Edition","Total Copies","Price (₹)","Shelf Location","Description"]
    col_widths = [28, 22, 20, 18, 22, 16, 12, 14, 13, 12, 16, 36]

    HEADER_FILL = PatternFill("solid", fgColor="0A1628")
    REQ_FILL    = PatternFill("solid", fgColor="1E3A5F")   # brighter for required cols
    HEADER_FONT = Font(name="Arial", color="FFFFFF", bold=True, size=10)
    NOTE_FONT   = Font(name="Arial", italic=True, size=8, color="6B7280")
    NOTE_FILL   = PatternFill("solid", fgColor="F3F4F6")
    SAMPLE_FONT = Font(name="Arial", size=9, color="374151")
    SAMPLE_FILL = PatternFill("solid", fgColor="EFF6FF")
    CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    THIN        = Side(style="thin", color="D1D5DB")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    REQUIRED    = {"Title", "Author", "ISBN", "Total Copies"}

    # ── Row 1: headers ────────────────────────────────────────────────
    ws.append(headers)
    ws.row_dimensions[1].height = 26
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill      = REQ_FILL if h in REQUIRED else HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Row 2: hint notes ─────────────────────────────────────────────
    notes = [
        "Required", "Required", "Required — unique per library",
        "e.g. Fiction", "e.g. Penguin", "e.g. 2020",
        "English / Bengali…", "e.g. 2nd",
        "Required — sets available copies", "e.g. 250.00",
        "e.g. A-shelf-3", "Optional summary",
    ]
    ws.append(notes)
    ws.row_dimensions[2].height = 18
    for ci, note in enumerate(notes, 1):
        cell = ws.cell(row=2, column=ci)
        cell.font      = NOTE_FONT
        cell.fill      = NOTE_FILL
        cell.alignment = CENTER
        cell.border    = BORDER

    # ── Row 3: sample data ────────────────────────────────────────────
    ws.append(["The Alchemist","Paulo Coelho","978-0062315007",
               "Fiction","HarperCollins",2014,"English","1st",5,350.00,"A-1",
               "A philosophical novel"])
    ws.row_dimensions[3].height = 18
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=ci)
        cell.font      = SAMPLE_FONT
        cell.fill      = SAMPLE_FILL
        cell.alignment = CENTER
        cell.border    = BORDER

    # ── Data validation ───────────────────────────────────────────────
    lang_dv = DataValidation(
        type="list", formula1='"English,Bengali,Hindi,Sanskrit,Nepali"',
        allow_blank=True, showDropDown=False,
    )
    lang_dv.sqref = "G4:G1000"
    ws.add_data_validation(lang_dv)

    copies_dv = DataValidation(
        type="whole", operator="greaterThanOrEqual", formula1="1",
        allow_blank=True, showErrorMessage=True,
        errorTitle="Invalid", error="Total Copies must be a whole number ≥ 1.",
    )
    copies_dv.sqref = "I4:I1000"
    ws.add_data_validation(copies_dv)

    price_dv = DataValidation(
        type="decimal", operator="greaterThanOrEqual", formula1="0",
        allow_blank=True, showErrorMessage=True,
        errorTitle="Invalid", error="Price must be a positive number (e.g. 250.00).",
    )
    price_dv.sqref = "J4:J1000"
    ws.add_data_validation(price_dv)

    ws.freeze_panes = "A4"   # keep header + notes rows visible while scrolling

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="book_import_template.xlsx"'
    return resp