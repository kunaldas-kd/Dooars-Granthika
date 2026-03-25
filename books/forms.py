from django import forms
from django.utils.text import slugify

from .models import Book, Category

# Accepted MIME types for cover images
ACCEPTED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/gif", "image/webp"}
MAX_IMAGE_BYTES = 5 * 1024 * 1024  # 5 MB


class BookForm(forms.ModelForm):

    # Replace the model's BinaryField with a plain file upload widget
    cover_image = forms.FileField(
        required=False,
        label="Cover Image",
        widget=forms.ClearableFileInput(attrs={"accept": "image/*"}),
        help_text="JPEG / PNG / WEBP / GIF — max 5 MB. Stored in the database.",
    )

    # Optional free-text field for creating a new category on the fly
    new_category = forms.CharField(
        required=False,
        max_length=100,
        widget=forms.TextInput(attrs={
            "placeholder": "e.g. Science Fiction",
            "autocomplete": "off",
            "id": "id_new_category",
        }),
        label="Or create a new category",
        help_text="Leave blank to use the dropdown above. "
                  "If a similar category already exists it will be reused automatically.",
    )

    class Meta:
        model  = Book
        # NOTE: book_id is intentionally omitted — it is auto-generated and editable=False.
        fields = [
            "title", "author", "isbn", "category",
            "publisher", "publication_year", "language", "edition",
            "total_copies", "shelf_location",
            "price",
            "description",
            # cover_image and new_category are declared above as extra fields
        ]
        widgets = {
            "title":            forms.TextInput(attrs={"placeholder": "e.g. The Great Gatsby"}),
            "author":           forms.TextInput(attrs={"placeholder": "e.g. F. Scott Fitzgerald"}),
            "isbn":             forms.TextInput(attrs={"placeholder": "e.g. 978-0-7432-7356-5"}),
            "publisher":        forms.TextInput(attrs={"placeholder": "e.g. Penguin Books"}),
            "publication_year": forms.NumberInput(attrs={"placeholder": "e.g. 2020", "min": 1000, "max": 2099}),
            "edition":          forms.TextInput(attrs={"placeholder": "e.g. 3rd Edition"}),
            "shelf_location":   forms.TextInput(attrs={"placeholder": "e.g. A-12-F"}),
            "total_copies":     forms.NumberInput(attrs={"placeholder": "e.g. 10", "min": 0}),
            "price":            forms.NumberInput(attrs={"placeholder": "e.g. 350.00", "min": 0, "step": "0.01", "required": True}),
            "description":      forms.Textarea(attrs={
                                    "rows": 4,
                                    "placeholder": "A brief description of the book…",
                                }),
        }

    # ------------------------------------------------------------------
    # Pass user= when constructing the form in the view
    # ------------------------------------------------------------------
    def __init__(self, *args, **kwargs):
        self.user = kwargs.pop("user", None)
        super().__init__(*args, **kwargs)

        if self.user is not None:
            self.fields["category"].queryset = Category.objects.filter(owner=self.user)
        self.fields["category"].required    = False
        self.fields["category"].empty_label = "— Select existing category —"

    # ------------------------------------------------------------------
    # Validation
    # ------------------------------------------------------------------
    def clean_cover_image(self):
        """Read uploaded file into bytes; validate type & size."""
        upload = self.cleaned_data.get("cover_image")

        # ClearableFileInput sends False when the user ticks "Clear"
        if upload is False:
            return False   # sentinel: caller will wipe the stored image

        if not upload:
            # No new file uploaded and no clear requested — keep existing
            return None

        # Size check
        if upload.size > MAX_IMAGE_BYTES:
            raise forms.ValidationError(
                f"Image too large ({upload.size // 1024} KB). Maximum is 5 MB."
            )

        # MIME type check
        mime = getattr(upload, "content_type", "")
        if mime not in ACCEPTED_IMAGE_TYPES:
            raise forms.ValidationError(
                f"Unsupported file type '{mime}'. "
                "Please upload a JPEG, PNG, WEBP, or GIF."
            )

        # Return a dict so the view can unpack both bytes and mime type
        return {"data": upload.read(), "mime": mime}

    def clean(self):
        cleaned      = super().clean()
        total        = cleaned.get("total_copies")
        new_cat_name = cleaned.get("new_category", "").strip()
        chosen_cat   = cleaned.get("category")

        # ── 1. Price required ─────────────────────────────────────────
        price = cleaned.get("price")
        if price is None:
            self.add_error("price", "Price is required.")
        elif price < 0:
            self.add_error("price", "Price cannot be negative.")

        # ── 2. Category required ──────────────────────────────────────
        if not chosen_cat and not new_cat_name:
            self.add_error(
                "category",
                "Please select an existing category or type a new one below.",
            )
            return cleaned

        # ── 3. Auto-create / reuse category ──────────────────────────
        if new_cat_name:
            qs = Category.objects.filter(owner=self.user, name__iexact=new_cat_name)
            if qs.exists():
                matched = qs.first()
                cleaned["category"] = matched
                self._reused_category = matched.name
            else:
                new_cat = Category.objects.create(
                    owner=self.user,
                    name=new_cat_name,
                    slug=slugify(new_cat_name),
                )
                cleaned["category"] = new_cat
                self._created_category = new_cat.name

        return cleaned


# ═══════════════════════════════════════════════════════════════
# Excel Import Form
# ═══════════════════════════════════════════════════════════════

EXCEL_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/vnd.ms-excel",                                            # .xls
    # browsers sometimes send these
    "application/octet-stream",
    "application/zip",
}
MAX_EXCEL_BYTES = 10 * 1024 * 1024  # 10 MB

# Expected column names (case-insensitive).  ISBN is the only hard-required one.
REQUIRED_COLS  = {"isbn"}
OPTIONAL_COLS  = {
    "title", "author", "category", "publisher",
    "publication year", "language", "edition",
    "shelf location", "total copies", "description",
    "price", "price (₹)",
}
ALL_COLS = REQUIRED_COLS | OPTIONAL_COLS


class ExcelImportForm(forms.Form):
    """Step-1 form: upload an .xlsx / .xls file for preview."""

    excel_file = forms.FileField(
        label="Excel File (.xlsx / .xls)",
        widget=forms.ClearableFileInput(attrs={"accept": ".xlsx,.xls"}),
        help_text="Max 10 MB. First row must be a header row.",
    )

    def clean_excel_file(self):
        upload = self.cleaned_data["excel_file"]

        if upload.size > MAX_EXCEL_BYTES:
            raise forms.ValidationError(
                f"File too large ({upload.size // 1024} KB). Maximum is 10 MB."
            )

        ext = upload.name.rsplit(".", 1)[-1].lower()
        if ext not in ("xlsx", "xls"):
            raise forms.ValidationError(
                "Only .xlsx and .xls files are accepted."
            )

        # Try to parse with openpyxl / xlrd to catch corrupt files early
        try:
            import openpyxl, io as _io
            data = upload.read()
            upload.seek(0)
            wb = openpyxl.load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else ""
                       for c in next(ws.iter_rows(min_row=1, max_row=1))]
            if not any(h in ALL_COLS for h in headers):
                raise forms.ValidationError(
                    "No recognised columns found. "
                    "Expected headers like: Title, Author, ISBN, Category…"
                )
        except forms.ValidationError:
            raise
        except Exception as e:
            raise forms.ValidationError(f"Could not read the file: {e}")

        upload.seek(0)
        return upload


def parse_excel_rows(file_obj, user):
    """
    Parse an uploaded Excel file and return a list of row-result dicts:

        {
          "row":    int,           # 1-based sheet row number of this data row
          "data":   dict,          # cleaned field values (category = Category instance)
          "status": "new" | "duplicate" | "error",
          "errors": [str, …],
          "book":   Book | None,   # existing Book instance if duplicate
        }

    Auto-creates categories (per user) when they don't exist yet.
    Does NOT save any Book records — that is the view's job after confirmation.
    """
    import io as _io
    import openpyxl
    from decimal import Decimal as _D, InvalidOperation
    from django.utils.text import slugify as _slugify

    # ── Column-name normaliser ────────────────────────────────────────
    # Maps any reasonable header spelling → internal snake_case key.
    _COL_MAP = {
        "title":            "title",
        "author":           "author",
        "isbn":             "isbn",
        "category":         "category",
        "publisher":        "publisher",
        "publication year": "publication_year",
        "publicationyear":  "publication_year",
        "publication_year": "publication_year",
        "language":         "language",
        "edition":          "edition",
        "total copies":     "total_copies",
        "totalcopies":      "total_copies",
        "total_copies":     "total_copies",
        "price (₹)":        "price",
        "price(₹)":         "price",
        "price (rs)":       "price",
        "price":            "price",
        "shelf location":   "shelf_location",
        "shelflocation":    "shelf_location",
        "shelf_location":   "shelf_location",
        "description":      "description",
        # legacy — silently ignored; value always derived from total_copies
        "available copies": None,
        "available_copies": None,
    }

    raw_data = file_obj.read()
    wb = openpyxl.load_workbook(_io.BytesIO(raw_data), read_only=True, data_only=True)

    # Accept sheet named "Books" or fall back to active sheet
    ws = wb["Books"] if "Books" in wb.sheetnames else wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return []

    # ── Find header row (skip blank / hint rows) ──────────────────────
    header_row_idx = None
    for i, row in enumerate(all_rows):
        normalised = [str(c).strip().lower() if c is not None else "" for c in row]
        if "isbn" in normalised or "title" in normalised:
            header_row_idx = i
            break
    if header_row_idx is None:
        return []

    raw_headers = all_rows[header_row_idx]
    # col_index → internal key  (None = ignore column)
    col_index_map = {}
    for ci, h in enumerate(raw_headers):
        key = _COL_MAP.get(str(h).strip().lower() if h is not None else "")
        if key:   # key is None for ignored cols and missing entries
            col_index_map[ci] = key

    results = []

    for row_idx, raw_row in enumerate(
        all_rows[header_row_idx + 1:],
        start=header_row_idx + 2,   # 1-based sheet row number
    ):
        # Skip entirely blank rows and the hint/notes row (row 2 in template)
        if all(v is None or str(v).strip() == "" for v in raw_row):
            continue

        # Build raw dict from mapped columns only
        row_dict = {
            key: (raw_row[ci] if ci < len(raw_row) else None)
            for ci, key in col_index_map.items()
        }

        errors   = []
        data_out = {}

        # ── ISBN (required) ───────────────────────────────────────────
        isbn_raw = str(row_dict.get("isbn") or "").strip()
        if not isbn_raw or isbn_raw.lower() == "none":
            errors.append("ISBN is required.")
            results.append({"row": row_idx, "data": {}, "status": "error",
                             "errors": errors, "book": None})
            continue
        data_out["isbn"] = isbn_raw

        # ── Title & Author (required) ─────────────────────────────────
        for req_field in ("title", "author"):
            val = str(row_dict.get(req_field) or "").strip()
            if not val or val.lower() == "none":
                errors.append(f"{req_field.title()} is required.")
            data_out[req_field] = val

        # ── Duplicate check (by ISBN) ─────────────────────────────────
        existing = Book.objects.filter(owner=user, isbn=isbn_raw).first()
        status   = "duplicate" if existing else "new"

        # ── Plain text fields ─────────────────────────────────────────
        for field in ("publisher", "edition", "shelf_location",
                      "language", "description"):
            val = row_dict.get(field)
            data_out[field] = str(val).strip() if val is not None else ""

        # ── publication_year ──────────────────────────────────────────
        yr = row_dict.get("publication_year")
        if yr is not None and str(yr).strip() not in ("", "None"):
            try:
                yr_int = int(float(str(yr)))
                if 1000 <= yr_int <= 2099:
                    data_out["publication_year"] = yr_int
                else:
                    errors.append(f"Publication Year '{yr}' is out of range 1000–2099.")
                    data_out["publication_year"] = None
            except (ValueError, TypeError):
                errors.append(f"Publication Year '{yr}' is not a valid number.")
                data_out["publication_year"] = None
        else:
            data_out["publication_year"] = None

        # ── total_copies ──────────────────────────────────────────────
        tc_val = row_dict.get("total_copies")
        if tc_val is not None and str(tc_val).strip() not in ("", "None"):
            try:
                tc_int = int(float(str(tc_val)))
                if tc_int < 1:
                    errors.append("Total Copies must be at least 1.")
                    data_out["total_copies"] = 1
                else:
                    data_out["total_copies"] = tc_int
            except (ValueError, TypeError):
                errors.append(f"Total Copies '{tc_val}' is not a valid number.")
                data_out["total_copies"] = 1
        else:
            data_out["total_copies"] = 1
        # available_copies is always derived — never read from the sheet
        data_out["available_copies"] = data_out["total_copies"]

        # ── price ─────────────────────────────────────────────────────
        price_val = row_dict.get("price")
        if price_val is None or str(price_val).strip() in ("", "None"):
            errors.append("Price is required.")
            data_out["price"] = None
        else:
            try:
                p = _D(str(price_val).replace(",", "").strip()).quantize(_D("0.01"))
                if p < 0:
                    errors.append("Price cannot be negative.")
                    data_out["price"] = None
                else:
                    data_out["price"] = p
            except (InvalidOperation, Exception):
                errors.append(f"Price '{price_val}' is not a valid number.")
                data_out["price"] = None

        # ── Category — auto-create if not found ───────────────────────
        cat_name = str(row_dict.get("category") or "").strip()
        if cat_name and cat_name.lower() != "none":
            cat_qs = Category.objects.filter(owner=user, name__iexact=cat_name)
            if cat_qs.exists():
                data_out["category"] = cat_qs.first()
                data_out["_category_created"] = False
            else:
                new_cat = Category.objects.create(
                    owner=user,
                    name=cat_name,
                    slug=_slugify(cat_name),
                )
                data_out["category"] = new_cat
                data_out["_category_created"] = True
        else:
            data_out["category"] = None
            data_out["_category_created"] = False

        results.append({
            "row":    row_idx,
            "data":   data_out,
            "status": "error" if errors else status,
            "errors": errors,
            "book":   existing,
        })

    return results